# -*- coding: utf-8 -*-
"""
Dựng báo cáo từ bond_issuance_raw.csv:
  1) TPDN_PhatHanh_TrongNuoc.xlsx  - Excel nhiều sheet + biểu đồ
  2) dashboard_data.json           - dữ liệu tổng hợp cho dashboard.html
Chạy sau bond_issuance_scraper.py
"""
import json
import os
import re
from datetime import datetime

import pandas as pd
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Phân loại ngành (13 nhóm: OVERRIDES tay -> VietCap ICB -> từ khóa) - dùng chung
from sector_map import classify, classify_with_source, order_groups, GROUP_ORDER
from vsd_xref import load_xref, doi_chieu, mo_ta_nguon

RAW = "bond_issuance_raw.csv"
BBRAW = "bond_buyback_raw.csv"
SECRAW = "bond_secondary_raw.csv"
CATRAW = "bond_catalog_raw.csv"   # danh mục ĐKGD: crosswalk mã giao dịch <-> mã CBTT <-> ISIN <-> TCPH
RATINGRAW = "bond_rating_raw.csv"  # xếp hạng tín nhiệm (bond_rating_scraper.py)
LPRAW = "bond_latepay_raw.csv"     # tin bất thường / chậm trả (bond_latepay_scraper.py)
XLSX = "TPDN_PhatHanh_TrongNuoc.xlsx"
TY = 1e9  # 1 tỷ đồng


def sanitize_nan(o):
    """NaN/Infinity -> None đệ quy. JSON chuẩn không có NaN; nếu để lọt, dashboard nhúng
    `const DATA={...}` sẽ đọc thành số NaN và làm hỏng thang đo mọi biểu đồ dùng Math.max."""
    if isinstance(o, dict):
        return {k: sanitize_nan(v) for k, v in o.items()}
    if isinstance(o, (list, tuple)):
        return [sanitize_nan(v) for v in o]
    if isinstance(o, float) and (o != o or o in (float("inf"), float("-inf"))):
        return None
    return o


# ---------- phân loại nhóm tổ chức phát hành ----------
# classify() import từ sector_map (14 nhóm: gán tay DN lớn + từ khóa).


def tenor_bucket(period: str, remain_days) -> str:
    # ưu tiên tính từ số ngày kỳ hạn gốc nếu parse được
    s = (period or "")
    m = re.search(r"(\d+)\s*(Năm|Tháng)", s)
    years = None
    if m:
        n = int(m.group(1))
        years = n if "Năm" in m.group(2) else n / 12.0
    if years is None:
        return "Không xác định"
    if years <= 1:
        return "≤ 1 năm"
    if years <= 3:
        return "> 1 - 3 năm"
    if years <= 5:
        return "> 3 - 5 năm"
    if years <= 10:
        return "> 5 - 10 năm"
    return "> 10 năm"


def rate_bucket(r) -> str:
    if pd.isna(r):
        return "Không rõ"
    if r < 6:
        return "< 6%"
    if r < 8:
        return "6 - 8%"
    if r < 10:
        return "8 - 10%"
    if r < 12:
        return "10 - 12%"
    return "≥ 12%"


def load():
    df = pd.read_csv(RAW)
    df["ph_date"] = pd.to_datetime(df["ngay_phat_hanh"], format="%d/%m/%Y", errors="coerce")
    df["post_date"] = pd.to_datetime(df["ngay_dang_tin"], format="%d/%m/%Y", errors="coerce")
    df = df.dropna(subset=["ph_date"])
    # Gộp các lần CÔNG BỐ LẠI của cùng một đợt phát hành (cùng mã TP + ngày phát hành):
    # giữ bản đăng tin mới nhất để tránh cộng trùng giá trị. Vẫn giữ các đợt khác ngày phát hành.
    n0 = len(df)
    df = df.sort_values("post_date").drop_duplicates(["ma_tp", "ngay_phat_hanh"], keep="last")
    print(f"Gộp trùng công bố: {n0} -> {len(df)} đợt phát hành ({df['ma_tp'].nunique()} mã TP)")
    df["nam"] = df["ph_date"].dt.year
    df["thang"] = df["ph_date"].dt.to_period("M").astype(str)
    # HNX có đợt công bố khối lượng = 0 -> gia_tri_phat_hanh rỗng. Bắt buộc coerce về 0:
    # để NaN thì nó lọt vào dashboard_data.json và làm NaN cả thang đo chart (chart trắng trơn).
    _gt_raw = pd.to_numeric(df["gia_tri_phat_hanh"], errors="coerce")
    _n_missing = int(_gt_raw.isna().sum())
    if _n_missing:
        print(f"CẢNH BÁO: {_n_missing} đợt phát hành không có giá trị (khối lượng = 0) -> tính = 0 tỷ")
    df["gia_tri_ty"] = _gt_raw.fillna(0.0) / TY
    _src = df["ten_dn"].apply(classify_with_source)
    df["nhom"] = _src.apply(lambda t: t[0])
    df["nhom_src"] = _src.apply(lambda t: t[1])   # nguồn ngành: Override / VietCap ICB / Từ khóa / Khác
    # tóm tắt nguồn phân ngành theo giá trị (minh bạch mức đóng góp của ICB VietCap)
    _sv = df.groupby("nhom_src")["gia_tri_ty"].sum()
    print("Nguồn phân ngành (tỷ): " + " · ".join(f"{k} {v:,.0f}" for k, v in _sv.sort_values(ascending=False).items()))
    df["ky_han_nhom"] = df.apply(lambda x: tenor_bucket(x["ky_han"], x["ky_han_con_lai"]), axis=1)
    df["ls_nhom"] = df["lai_suat_num"].apply(rate_bucket)
    # coupon bình quân gia quyền theo giá trị
    return df


# ---------- Excel ----------
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def write_df(ws, df, number_cols=None, pct_cols=None):
    number_cols = number_cols or []
    pct_cols = pct_cols or []
    ws.append(list(df.columns))
    for _, r in df.iterrows():
        ws.append(list(r.values))
    style_header(ws, len(df.columns))
    for i, col in enumerate(df.columns, 1):
        letter = get_column_letter(i)
        width = max(12, min(48, int(df[col].astype(str).str.len().max() if len(df) else 12) + 2))
        ws.column_dimensions[letter].width = width
        if col in number_cols:
            for row in range(2, len(df) + 2):
                ws.cell(row=row, column=i).number_format = "#,##0"
        if col in pct_cols:
            for row in range(2, len(df) + 2):
                ws.cell(row=row, column=i).number_format = "0.0%"
    ws.freeze_panes = "A2"


def wavg(g):
    v = g["gia_tri_phat_hanh"]
    r = g["lai_suat_num"]
    mask = r.notna()
    if v[mask].sum() == 0:
        return None
    return (v[mask] * r[mask]).sum() / v[mask].sum()


def build_excel(df, out=None):
    from openpyxl import Workbook
    wb = Workbook()

    # --- Sheet 1: Tổng quan (KPI) ---
    ws = wb.active
    ws.title = "Tổng quan"
    total_val = df["gia_tri_ty"].sum()
    now = datetime.now()
    ytd = df[df["nam"] == df["nam"].max()]
    kpis = [
        ["CHỈ TIÊU", "GIÁ TRỊ"],
        ["Số đợt phát hành", len(df)],
        ["Số mã trái phiếu", df["ma_tp"].nunique()],
        ["Số tổ chức phát hành", df["ten_dn"].nunique()],
        ["Tổng giá trị phát hành (tỷ VNĐ)", round(total_val, 0)],
        ["Giá trị bình quân/đợt (tỷ VNĐ)", round(total_val / len(df), 1)],
        ["Coupon bình quân gia quyền (%)", round(wavg(df), 2)],
        [f"Số đợt năm {int(df['nam'].max())}", len(ytd)],
        [f"Giá trị năm {int(df['nam'].max())} (tỷ VNĐ)", round(ytd['gia_tri_ty'].sum(), 0)],
        ["Khoảng thời gian dữ liệu", f"{df['ph_date'].min():%d/%m/%Y} - {df['ph_date'].max():%d/%m/%Y}"],
        ["Cập nhật lúc", now.strftime("%d/%m/%Y %H:%M")],
    ]
    for r in kpis:
        ws.append(r)
    style_header(ws, 2)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 26
    for row in range(2, len(kpis) + 1):
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
        if isinstance(ws.cell(row=row, column=2).value, (int, float)):
            ws.cell(row=row, column=2).number_format = "#,##0.0"
    ws.append([])
    ws.append(["Nguồn: cbonds.hnx.vn - Kết quả chào bán TPDN riêng lẻ trong nước (HNX/CBIS)"])

    # --- Sheet 2: Theo năm ---
    by_year = df.groupby("nam").apply(
        lambda g: pd.Series({
            "So dot": len(g),
            "Gia tri (ty VND)": round(g["gia_tri_ty"].sum(), 0),
            "Coupon BQ (%)": round(wavg(g) or 0, 2),
        }), include_groups=False).reset_index().rename(columns={"nam": "Nam"})
    ws2 = wb.create_sheet("Theo năm")
    write_df(ws2, by_year, number_cols=["So dot", "Gia tri (ty VND)"])
    # chart
    ch = BarChart(); ch.type = "col"; ch.title = "Giá trị phát hành theo năm (tỷ VNĐ)"
    data = Reference(ws2, min_col=3, min_row=1, max_row=len(by_year) + 1)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=len(by_year) + 1)
    ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
    ch.height = 8; ch.width = 16
    ws2.add_chart(ch, "F2")

    # --- Sheet 3: Theo tháng ---
    by_month = df.groupby("thang").apply(
        lambda g: pd.Series({
            "So dot": len(g),
            "Gia tri (ty VND)": round(g["gia_tri_ty"].sum(), 0),
            "Coupon BQ (%)": round(wavg(g) or 0, 2),
        }), include_groups=False).reset_index().rename(columns={"thang": "Thang"})
    ws3 = wb.create_sheet("Theo tháng")
    write_df(ws3, by_month, number_cols=["So dot", "Gia tri (ty VND)"])
    lc = LineChart(); lc.title = "Giá trị phát hành theo tháng (tỷ VNĐ)"
    data = Reference(ws3, min_col=3, min_row=1, max_row=len(by_month) + 1)
    cats = Reference(ws3, min_col=1, min_row=2, max_row=len(by_month) + 1)
    lc.add_data(data, titles_from_data=True); lc.set_categories(cats)
    lc.height = 8; lc.width = 24
    ws3.add_chart(lc, "F2")

    # --- Sheet 4: Theo nhóm ngành ---
    by_grp = df.groupby("nhom").apply(
        lambda g: pd.Series({
            "So dot": len(g),
            "Gia tri (ty VND)": round(g["gia_tri_ty"].sum(), 0),
            "Ty trong": g["gia_tri_ty"].sum(),
            "Coupon BQ (%)": round(wavg(g) or 0, 2),
        }), include_groups=False).reset_index().rename(columns={"nhom": "Nhom TCPH"})
    by_grp["Ty trong"] = by_grp["Ty trong"] / by_grp["Ty trong"].sum()
    by_grp = by_grp.sort_values("Gia tri (ty VND)", ascending=False)
    ws4 = wb.create_sheet("Theo nhóm ngành")
    write_df(ws4, by_grp, number_cols=["So dot", "Gia tri (ty VND)"], pct_cols=["Ty trong"])

    # --- Sheet 5: Theo kỳ hạn & lãi suất ---
    by_ten = df.groupby("ky_han_nhom").apply(
        lambda g: pd.Series({"So dot": len(g), "Gia tri (ty VND)": round(g["gia_tri_ty"].sum(), 0)}),
        include_groups=False).reset_index().rename(columns={"ky_han_nhom": "Ky han"})
    by_rate = df.groupby("ls_nhom").apply(
        lambda g: pd.Series({"So dot": len(g), "Gia tri (ty VND)": round(g["gia_tri_ty"].sum(), 0)}),
        include_groups=False).reset_index().rename(columns={"ls_nhom": "Nhom lai suat"})
    ws5 = wb.create_sheet("Kỳ hạn & Lãi suất")
    write_df(ws5, by_ten, number_cols=["So dot", "Gia tri (ty VND)"])
    start = len(by_ten) + 4
    ws5.cell(row=start - 1, column=1, value="Phân bố theo nhóm lãi suất").font = Font(bold=True)
    ws5.append([])
    r0 = ws5.max_row + 1
    ws5.cell(row=r0, column=1, value="Nhom lai suat"); ws5.cell(row=r0, column=2, value="So dot")
    ws5.cell(row=r0, column=3, value="Gia tri (ty VND)")
    style_header(ws5, 3, row=r0)
    for _, r in by_rate.iterrows():
        ws5.append([r["Nhom lai suat"], r["So dot"], r["Gia tri (ty VND)"]])

    # --- Sheet 6: Top tổ chức phát hành ---
    by_iss = df.groupby("ten_dn").apply(
        lambda g: pd.Series({
            "So dot": len(g),
            "Gia tri (ty VND)": round(g["gia_tri_ty"].sum(), 0),
            "Coupon BQ (%)": round(wavg(g) or 0, 2),
        }), include_groups=False).reset_index().rename(columns={"ten_dn": "To chuc phat hanh"})
    by_iss = by_iss.sort_values("Gia tri (ty VND)", ascending=False).head(30)
    ws6 = wb.create_sheet("Top TCPH")
    write_df(ws6, by_iss, number_cols=["So dot", "Gia tri (ty VND)"])

    # --- Sheet 7: Dữ liệu thô ---
    raw_cols = ["ngay_dang_tin", "ten_dn", "ma_tp", "nhom", "ky_han", "ngay_phat_hanh",
                "ngay_dao_han", "khoi_luong_num", "menh_gia_num", "gia_tri_ty",
                "loai_lai_suat", "lai_suat_num", "mua_lai_hoan_doi", "tinh_trang"]
    raw = df[raw_cols].copy()
    raw.columns = ["Ngày đăng tin", "Tổ chức phát hành", "Mã TP", "Nhóm", "Kỳ hạn",
                   "Ngày phát hành", "Ngày đáo hạn", "Khối lượng", "Mệnh giá",
                   "Giá trị (tỷ VNĐ)", "Loại lãi suất", "Lãi suất (%)",
                   "Mua lại/Hoán đổi", "Tình trạng"]
    ws7 = wb.create_sheet("Dữ liệu thô")
    write_df(ws7, raw, number_cols=["Khối lượng", "Mệnh giá", "Giá trị (tỷ VNĐ)"])

    if out is not None:
        build_outstanding_sheets(wb, out)

    wb.save(XLSX)
    print(f"Đã tạo {XLSX} ({len(wb.sheetnames)} sheet)")


# ---------- Mua lại & Dư nợ đang lưu hành ----------
def load_buyback():
    bb = pd.read_csv(BBRAW)
    bb["dtm"] = pd.to_datetime(bb["ngay_mua_lai"], format="%d/%m/%Y", errors="coerce")
    bb["nhom"] = bb["ten_dn"].apply(classify)
    return bb


def load_secondary(df, bb=None):
    """Đọc giao dịch THỨ CẤP (bond_secondary_raw.csv, theo mã×ngày có GD) -> block cho dashboard:
       rows cấp (mã, ngày) + meta mã->[tên DN, nhóm] + KPI toàn kỳ. Trả None nếu chưa có file."""
    if not os.path.exists(SECRAW):
        return None
    sec = pd.read_csv(SECRAW)
    sec["dtm"] = pd.to_datetime(sec["ngay_gd"], format="%d/%m/%Y", errors="coerce")
    sec = sec.dropna(subset=["dtm"])
    sec["gt"] = pd.to_numeric(sec["gt_num"], errors="coerce").fillna(0.0)          # đồng
    sec["kl"] = pd.to_numeric(sec["kl_num"], errors="coerce").fillna(0).astype("int64")
    sec = sec[sec["gt"] > 0]
    if sec.empty:
        return None
    # MAP mã GIAO DỊCH -> TCPH. Mã giao dịch thứ cấp (vd VJC12101) KHÁC mã phát hành/CBTT (vd VJCH2101),
    # nên KHÔNG join trực tiếp với bond_issuance_raw. Dùng DANH MỤC ĐKGD (bond_catalog_raw.csv) làm
    # crosswalk chuẩn: ma_gd -> ma_cbtt/ISIN/tên TCPH (phủ 100% mã có giao dịch).
    cat_name, cat_cbtt, cat_isin = {}, {}, {}
    if os.path.exists(CATRAW):
        cat = pd.read_csv(CATRAW, dtype=str).fillna("")
        for r in cat.itertuples(index=False):
            g = (r.ma_gd or "").strip()
            if g and g not in cat_name:
                cat_name[g] = (r.ten_tcph or "").strip()
                cat_cbtt[g] = (r.ma_cbtt or "").strip()
                cat_isin[g] = (r.isin or "").strip()
    else:
        print(f"CẢNH BÁO: thiếu {CATRAW} -> mã giao dịch sẽ không map được TCPH. "
              f"Chạy: python bond_catalog_scraper.py")
    # dự phòng: nếu 1 mã không có trong danh mục, thử bảng phát hành/mua lại
    fb_name = dict(zip(df["ma_tp"], df["ten_dn"]))
    if bb is not None:
        for ma, dn in zip(bb["ma_tp"], bb["ten_dn"]):
            fb_name.setdefault(ma, dn)
    sec["dn"] = sec["ma_tp"].map(lambda m: cat_name.get(m) or fb_name.get(m, "")).fillna("")
    sec["cbtt"] = sec["ma_tp"].map(lambda m: cat_cbtt.get(m, ""))
    sec["nhom"] = sec["dn"].apply(lambda n: classify(n) if n else "Khác")
    n_unmapped = int((sec["dn"] == "").groupby(sec["ma_tp"]).first().sum())
    print(f"  map mã giao dịch: {sec['ma_tp'].nunique()} mã · chưa map {n_unmapped} mã (nhóm 'Khác')")

    # rows cấp (mã, ngày) - gt quy về tỷ để nhẹ payload; dashboard tự lọc & tổng hợp theo ngày/tuần/tháng/quý/năm
    sec_rows = [{"d": r.ngay_gd, "ma": r.ma_tp, "gt": round(r.gt / TY, 3), "kl": int(r.kl)}
                for r in sec.itertuples(index=False)]
    # meta mã -> [tên DN, nhóm, mã CBTT] (khử trùng) - mã CBTT để đối chiếu với dữ liệu phát hành
    meta = {}
    for r in sec.itertuples(index=False):
        if r.ma_tp not in meta:
            meta[r.ma_tp] = [r.dn, r.nhom, r.cbtt]

    by_day = sec.groupby("ngay_gd")["gt"].sum()
    kpi = {
        "total_gt": round(sec["gt"].sum() / TY, 0),
        "total_kl": int(sec["kl"].sum()),
        "n_sessions": int(sec["ngay_gd"].nunique()),
        "n_bonds": int(sec["ma_tp"].nunique()),
        "avg_session": round(sec["gt"].sum() / TY / max(int(sec["ngay_gd"].nunique()), 1), 1),
        "max_session": round(by_day.max() / TY, 1),
        "max_day": by_day.idxmax(),
        "first": f"{sec['dtm'].min():%d/%m/%Y}",
        "last": f"{sec['dtm'].max():%d/%m/%Y}",
    }
    print(f"Giao dịch thứ cấp: {len(sec_rows):,} bản ghi (mã×ngày) · {kpi['n_sessions']} phiên · "
          f"{kpi['total_gt']:,.0f} tỷ · BQ {kpi['avg_session']:,.1f} tỷ/phiên")
    return {"rows": sec_rows, "meta": meta, "kpi": kpi}


def _short_agency(s):
    """Rút gọn tên đơn vị XHTN: lấy tên trong ngoặc nếu có (vd '...(FiinRatings)' -> 'FiinRatings')."""
    m = re.search(r"\(([^)]+)\)\s*$", (s or "").strip())
    return m.group(1).strip() if m else (s or "").strip()


def load_rating():
    """Đọc XẾP HẠNG TÍN NHIỆM (bond_rating_raw.csv) -> block cho dashboard.
       rows cấp kết quả XHTN + nhóm ngành TCPH + KPI. Trả None nếu chưa có file."""
    if not os.path.exists(RATINGRAW):
        return None
    r = pd.read_csv(RATINGRAW, dtype=str).fillna("")
    if r.empty:
        return None
    r["nhom"] = r["ten_tcph"].apply(classify)
    r["dv"] = r["don_vi_xhtn"].apply(_short_agency)
    # đối tượng XHTN: 'tp' = Trái phiếu (có mã TP) · 'org' = Tổ chức phát hành
    r["doi"] = r["doi_tuong_xhtn"].apply(lambda s: "tp" if "trái phiếu" in (s or "").lower() else "org")
    rows = [{"dv": x.dv, "dvfull": x.don_vi_xhtn, "doi": x.doi, "dn": x.ten_tcph, "ma": x.ma_tp,
             "kq": x.ket_qua_xhtn, "hl": x.hieu_luc_tu_ngay,
             "loai": x.loai_xep_hang, "nhom": x.nhom, "file": x.file_id}
            for x in r.itertuples(index=False)]
    is_org = r["doi"] == "org"
    kpi = {"n": int(len(r)), "n_tcph": int(r.loc[is_org, "ten_tcph"].nunique()),
           "n_org": int(is_org.sum()), "n_tp": int((~is_org).sum()),
           "n_dv": int(r["dv"].nunique())}
    print(f"Xếp hạng tín nhiệm: {kpi['n']} kết quả ({kpi['n_org']} TCPH · {kpi['n_tp']} trái phiếu) · "
          f"{kpi['n_tcph']} TCPH được xếp hạng · {kpi['n_dv']} đơn vị XHTN")
    return {"rows": rows, "kpi": kpi}


def _late_type(title):
    """Loại chậm trả từ tiêu đề CBTT: 'goc' (gốc) · 'lai' (lãi) · 'ca_hai' (cả gốc & lãi) · 'khac'."""
    import unicodedata
    t = "".join(c for c in unicodedata.normalize("NFD", title or "")
                if unicodedata.category(c) != "Mn").lower()
    goc, lai = ("goc" in t), ("lai" in t)
    if goc and lai:
        return "ca_hai"
    if goc:
        return "goc"
    if lai:
        return "lai"
    return "khac"


def _split_codes(s):
    return [c.strip() for c in str(s or "").split(",") if c.strip()]


def load_giahan_codes():
    """Tập mã TP có CBTT GIA HẠN / đổi điều khoản-kỳ hạn (bond_latepay_scraper.classify_event).
    Dùng làm BẰNG CHỨNG của HNX để cho phép dời ngày đáo hạn theo VSD."""
    if not os.path.exists(LPRAW):
        return set()
    lp = pd.read_csv(LPRAW, dtype=str).fillna("")
    if "loai_su_kien" not in lp.columns:
        return set()
    out = set()
    for s in lp[lp["loai_su_kien"] == "gia_han"]["ma_tp"]:
        out.update(c.upper() for c in _split_codes(s))
    return out


def outstanding_by_bond(df, bb):
    """Dư nợ gốc CÒN LẠI theo mã TP (đồng) = giá trị phát hành − đã mua lại, KHÔNG trừ đáo hạn
       (mã chậm trả tuy quá hạn danh nghĩa nhưng gốc thực tế CHƯA trả). Dùng để quy dư nợ cho mã chậm trả."""
    dno = {}
    if df is not None:
        for ma, v in df.groupby("ma_tp")["gia_tri_phat_hanh"].sum().items():
            if pd.notna(v):
                dno[ma] = float(v)
    if bb is not None and len(bb):
        latest = (bb.dropna(subset=["dtm"]).sort_values("dtm")
                    .drop_duplicates("ma_tp", keep="last"))
        for r in latest.itertuples(index=False):
            conlai = getattr(r, "gt_con_lai_num", None)
            face = getattr(r, "gt_phat_hanh_num", None)
            if pd.notna(conlai):
                dno[r.ma_tp] = float(conlai)          # còn lại sau mua lại (bản mới nhất)
            elif r.ma_tp not in dno and pd.notna(face):
                dno[r.ma_tp] = float(face)
    return dno


NEWSRAW = "bond_news_raw.csv"


def load_news():
    """Tab "TIN KHÁC" của trang Công bố thông tin (bond_news_scraper.py).

    LÝ DO CÓ MODULE NÀY (khảo sát 22/07/2026): doanh nghiệp công bố việc ĐÃ TRẢ NỢ ở tab
    "Tin khác", KHÔNG phải tab "Tin bất thường" mà ta vẫn quét -> dashboard chỉ thấy 5 lượt
    khắc phục trong khi thực tế có ~90, và 33 mã bị treo trạng thái "đang chậm trả" oan.

    Trả {"cure": {mã: {d, td, loai, dt}}} — với mỗi mã giữ sự kiện thanh toán MỚI NHẤT.
    Không xét ở đây việc sự kiện có sau lần chậm trả cuối hay không: so sánh ngày thuộc về
    nơi dựng danh sách chậm trả (dashboard), vì danh sách đó còn chịu bộ lọc thời gian.
    """
    if not os.path.exists(NEWSRAW):
        return None
    nw = pd.read_csv(NEWSRAW, dtype=str).fillna("")
    nw["dt"] = pd.to_datetime(nw["ngay_dang_tin"], format="%d/%m/%Y", errors="coerce")
    nw = nw.dropna(subset=["dt"]).sort_values("dt")
    cure = {}
    # tt_bo_sung = trả BỔ SUNG phần còn thiếu -> khắc phục MỘT PHẦN, KHÔNG kết luận hết chậm;
    # giữ riêng loại để dashboard hiển thị đúng mức độ, không gộp chung với 'khac_phuc'.
    for x in nw[nw["loai_su_kien"].isin(["khac_phuc", "tt_bo_sung", "tat_toan"])].itertuples(index=False):
        for c in _split_codes(x.ma_tp):
            cure[c] = {"d": x.ngay_dang_tin, "td": x.tieu_de, "loai": x.loai_su_kien}
    n_by = pd.Series([v["loai"] for v in cure.values()]).value_counts().to_dict() if cure else {}
    print(f"Tin khác: {len(nw)} tin · {len(cure)} mã có CBTT đã thanh toán "
          + " · ".join(f"{k} {v}" for k, v in n_by.items()))
    return {"cure": cure, "n_tin": int(len(nw))}


def load_latepay(dno_map=None, total_out_ty=None, news=None):
    """Đọc TIN BẤT THƯỜNG (bond_latepay_raw.csv), chỉ giữ sự kiện chậm trả/khắc phục -> block dashboard.
       rows cấp CBTT (+ loại chậm gốc/lãi/cả hai) + nhóm ngành + dư nợ theo mã + KPI (gồm % dư nợ chậm/tổng).
       Trả None nếu chưa có file / không có sự kiện chậm trả."""
    if not os.path.exists(LPRAW):
        return None
    lp = pd.read_csv(LPRAW, dtype=str).fillna("")
    lp = lp[lp["loai_su_kien"] != ""].copy()
    if lp.empty:
        return None
    lp["nhom"] = lp["ten_dn"].apply(classify)
    lp["lct"] = lp["tieu_de"].apply(_late_type)
    lp["dt"] = pd.to_datetime(lp["ngay_dang_tin"], format="%d/%m/%Y", errors="coerce")
    lp = lp.dropna(subset=["dt"]).sort_values("dt")
    rows = [{"d": x.ngay_dang_tin, "dn": x.ten_dn, "ma": x.ma_tp, "td": x.tieu_de,
             "loai": x.loai_su_kien, "lct": x.lct, "nhom": x.nhom, "file": x.file_id}
            for x in lp.itertuples(index=False)]
    cham = lp[lp["loai_su_kien"] == "cham_tra"]

    # dư nợ (tỷ) theo mã chậm trả + loại chậm gộp theo mã (goc/lai/ca_hai)
    dno_map = dno_map or {}
    code_types, code_dno = {}, {}
    for x in cham.itertuples(index=False):
        for c in _split_codes(x.ma_tp):
            code_types.setdefault(c, set()).add(x.lct)
            if c in dno_map:
                code_dno[c] = round(dno_map[c] / TY, 1)
    codes = list(code_types.keys())
    codes_goc = [c for c in codes if code_types[c] & {"goc", "ca_hai"}]
    dno_total = round(sum(code_dno.get(c, 0.0) for c in codes), 0)
    dno_goc = round(sum(code_dno.get(c, 0.0) for c in codes_goc), 0)

    kpi = {"n_events": int(len(cham)), "n_dn": int(cham["ten_dn"].nunique()),
           "n_ma": len(codes), "n_ma_dno": len(code_dno),
           "n_cured": int((lp["loai_su_kien"] == "khac_phuc").sum()),
           "dno_total": dno_total, "dno_goc": dno_goc,
           "total_out_ty": (round(total_out_ty, 0) if total_out_ty else None),
           "pct_total": (round(dno_total / total_out_ty * 100, 2) if total_out_ty else None),
           "pct_goc": (round(dno_goc / total_out_ty * 100, 2) if total_out_ty else None),
           "first": (f"{cham['dt'].min():%d/%m/%Y}" if len(cham) else ""),
           "last": (f"{cham['dt'].max():%d/%m/%Y}" if len(cham) else "")}
    # KÊNH KHẮC PHỤC THỨ HAI (tab "Tin khác") — chỉ giữ mã THỰC SỰ nằm trong danh sách chậm
    # trả, khỏi đẩy sang dashboard 183 mã không liên quan.
    cure = {c: v for c, v in ((news or {}).get("cure") or {}).items() if c in code_types}
    kpi["n_cure_news"] = len(cure)
    print(f"Chậm trả gốc/lãi: {kpi['n_events']} lượt · {kpi['n_dn']} TCPH · {kpi['n_ma']} mã "
          f"(khớp dư nợ {kpi['n_ma_dno']}) · dư nợ chậm {dno_total:,.0f} tỷ"
          + (f" = {kpi['pct_total']}% tổng dư nợ" if kpi["pct_total"] else "")
          + f" · {kpi['n_cured']} khắc phục (Tin bất thường)"
          + (f" + {len(cure)} mã có CBTT đã trả ở Tin khác" if cure else ""))
    return {"rows": rows, "kpi": kpi, "dno": code_dno, "cure": cure}


def load_giahan(out=None):
    """Tab GIA HẠN / ĐỔI ĐIỀU KHOẢN (user chốt 17/07/2026 — kênh này trước đây bị bỏ không dùng).

    Ý nghĩa: gia hạn = TÁI CẤU TRÚC NỢ -> tín hiệu rủi ro sớm, thường đi trước chậm trả.
    Dữ liệu: bond_latepay_raw.csv, loai_su_kien == 'gia_han' (classify_event của scraper).
    Gộp về CẤP MÃ: số lần gia hạn, lần gần nhất, dư nợ hiện tại, có kèm chậm trả không,
    ngày đáo hạn đã được dời theo VSD hay chưa.
    """
    if not os.path.exists(LPRAW):
        return None
    lp = pd.read_csv(LPRAW, dtype=str).fillna("")
    if "loai_su_kien" not in lp.columns:
        return None
    gh = lp[lp["loai_su_kien"] == "gia_han"].copy()
    if gh.empty:
        return None
    gh["dt"] = pd.to_datetime(gh["ngay_dang_tin"], format="%d/%m/%Y", errors="coerce")
    gh = gh.dropna(subset=["dt"]).sort_values("dt")

    cham_codes = set()
    for s in lp[lp["loai_su_kien"] == "cham_tra"]["ma_tp"]:
        cham_codes.update(c.upper() for c in _split_codes(s))

    # CHUỖI CẢNH BÁO 3 TẦNG (bổ sung 22/07/2026): luật buộc lấy ý kiến / họp người sở hữu
    # trái phiếu TRƯỚC khi gia hạn, mà gia hạn lại thường đi TRƯỚC chậm trả. Mốc "xin ý kiến"
    # vì thế là mắt xích SỚM NHẤT quan sát được của chuỗi tái cơ cấu nợ:
    #     xin ý kiến trái chủ  ->  gia hạn / đổi điều khoản  ->  chậm trả
    # Dữ liệu nằm sẵn trong 1.250 tin bất thường trước đây bỏ trống, không tốn thêm request.
    yk = lp[lp["loai_su_kien"] == "y_kien_trai_chu"].copy()
    yk["dt"] = pd.to_datetime(yk["ngay_dang_tin"], format="%d/%m/%Y", errors="coerce")
    yk = yk.dropna(subset=["dt"]).sort_values("dt")
    yk_first = {}                                  # mã -> LẦN ĐẦU xin ý kiến
    for x in yk.itertuples(index=False):
        for c in _split_codes(x.ma_tp):
            yk_first.setdefault(c.upper(), x.ngay_dang_tin)
    xp_codes = set()                               # mã của TCPH bị xử phạt hành chính
    for s in lp[lp["loai_su_kien"] == "xu_phat"]["ma_tp"]:
        xp_codes.update(c.upper() for c in _split_codes(s))

    # thông tin cấp mã từ bảng dư nợ (dư nợ, TCPH, nhóm, ĐH hiện hành, nguồn ĐH)
    info = {}
    if out is not None:
        for x in out["detail"].itertuples(index=False):
            info[str(x.ma_tp).upper()] = {
                "dn": x.dn, "nhom": x.nhom, "gt": round(x.remaining / TY, 1),
                "dh": (x.dh.strftime("%d/%m/%Y") if pd.notna(x.dh) else ""),
                "tt": x.trang_thai,
                "dh_goc": (x.dh.strftime("%d/%m/%Y") if pd.notna(x.dh) else ""),
                "dh_gh": (x.dh_gh.strftime("%d/%m/%Y") if pd.notna(x.dh_gh) else ""),
                "dh_gh_dt": (x.dh_gh if pd.notna(x.dh_gh) else None),
                "dh_dt": (x.dh if pd.notna(x.dh) else None)}

    agg = {}
    for x in gh.itertuples(index=False):
        for c in _split_codes(x.ma_tp):
            k = c.upper()
            a = agg.setdefault(k, {"ma": c, "n": 0, "lan_dau": x.ngay_dang_tin,
                                   "lan_cuoi": "", "dn_cbtt": x.ten_dn})
            a["n"] += 1
            a["lan_cuoi"] = x.ngay_dang_tin          # đã sort theo ngày -> giữ bản cuối
    today = pd.Timestamp(datetime.now().date())
    rows = []
    for k, a in agg.items():
        i = info.get(k, {})
        # HẠN CHÓT PHÁP LÝ = đáo hạn GỐC + 2 năm (NĐ08/2023: gia hạn TỐI ĐA 2 năm).
        # Quá mốc này thì hết đường gia hạn -> phải trả hoặc thành chậm trả (user chốt 17/07/2026).
        dg = i.get("dh_dt")
        hc = (dg + pd.Timedelta(days=730)) if dg is not None else None
        dm = i.get("dh_gh_dt")     # ngày đáo hạn theo LỊCH GIA HẠN MỚI NHẤT (VSD)
        rows.append({"ma": a["ma"], "n": a["n"], "d": a["lan_cuoi"],
                     "dn": i.get("dn") or a["dn_cbtt"], "nhom": i.get("nhom", "Khác"),
                     "gt": i.get("gt", None),
                     "dhg": i.get("dh_goc", ""),      # đáo hạn GỐC (bảng chính vẫn dùng ngày này)
                     "dhm": i.get("dh_gh", ""),       # đáo hạn theo lịch gia hạn MỚI NHẤT
                     "sn": (int((dm - dg).days) if (dm is not None and dg is not None) else None),
                     "hc": (hc.strftime("%d/%m/%Y") if hc is not None else ""),
                     "hcq": (bool(hc < today) if hc is not None else None),   # hạn chót đã qua?
                     "tt": i.get("tt", ""), "cham": (k in cham_codes),
                     "yk": yk_first.get(k, ""),          # ngày xin ý kiến trái chủ LẦN ĐẦU
                     # số ngày từ khi xin ý kiến đến khi CBTT gia hạn -> đo "báo trước bao lâu"
                     "yks": ((pd.to_datetime(a["lan_dau"], format="%d/%m/%Y", errors="coerce")
                              - pd.to_datetime(yk_first[k], format="%d/%m/%Y", errors="coerce")).days
                             if k in yk_first else None),
                     "xp": (k in xp_codes)})
    rows.sort(key=lambda r: (-(r["gt"] or 0), -r["n"]))

    n_dn = len({r["dn"] for r in rows})
    dno = round(sum(r["gt"] or 0 for r in rows), 0)
    n_cham = sum(1 for r in rows if r["cham"])
    n_moi = sum(1 for r in rows if r["dhm"])
    n_qua = sum(1 for r in rows if r["hcq"])
    n_kich = sum(1 for r in rows if r["sn"] and r["sn"] >= 728)
    n_yk = sum(1 for r in rows if r["yk"])
    n_xp = sum(1 for r in rows if r["xp"])
    _lead = [r["yks"] for r in rows if r["yks"] is not None and r["yks"] >= 0]
    lead_med = int(pd.Series(_lead).median()) if _lead else None
    kpi = {"n_events": int(len(gh)), "n_ma": len(rows), "n_dn": n_dn, "dno": dno,
           "n_cham": n_cham, "n_nhieu_lan": sum(1 for r in rows if r["n"] > 1),
           "n_moi": n_moi, "n_qua_hanchot": n_qua, "n_kich_tran": n_kich,
           "pct_cham": (round(n_cham / len(rows) * 100, 1) if rows else 0),
           "n_yk": n_yk, "n_xp": n_xp, "lead_med": lead_med,
           "n_yk_events": int(len(yk)),
           "first": f"{gh['dt'].min():%d/%m/%Y}", "last": f"{gh['dt'].max():%d/%m/%Y}"}
    print(f"Gia hạn/đổi điều khoản: {kpi['n_events']} lượt · {kpi['n_ma']} mã · {n_dn} TCPH "
          f"· dư nợ {dno:,.0f} tỷ · {n_cham} mã ({kpi['pct_cham']}%) kèm chậm trả "
          f"· {n_moi} mã có lịch gia hạn mới ({n_kich} kịch trần 2 năm) "
          f"· {n_qua} mã QUÁ hạn chót pháp lý"
          + (f" · {n_yk} mã có xin ý kiến trái chủ trước"
             + (f" (báo trước trung vị {lead_med} ngày)" if lead_med is not None else "") if n_yk else "")
          + (f" · {n_xp} mã của TCPH bị xử phạt" if n_xp else ""))
    # ev = cấp LƯỢT CBTT (cho biểu đồ theo thời gian); rows = cấp MÃ (cho bảng)
    ev = [{"d": x.ngay_dang_tin, "dn": x.ten_dn, "ma": x.ma_tp, "td": x.tieu_de,
           "nhom": classify(x.ten_dn), "file": x.file_id} for x in gh.itertuples(index=False)]
    return {"rows": rows, "ev": ev, "kpi": kpi}


def load_updates(max_changes=300):
    """Đọc trạng thái cập nhật (update_state.json) + nhật ký thay đổi gần đây (changes_log.csv)
       cho tab 'Cập nhật & thay đổi'. Trả None nếu chưa chạy update_daily lần nào."""
    state_path, log_path = "update_state.json", "changes_log.csv"
    if not os.path.exists(state_path) and not os.path.exists(log_path):
        return None
    state = None
    if os.path.exists(state_path):
        try:
            state = json.load(open(state_path, encoding="utf-8"))
        except Exception:
            state = None
    changes = []
    if os.path.exists(log_path):
        try:
            cl = pd.read_csv(log_path, dtype=str).fillna("")
            cl = cl.tail(max_changes).iloc[::-1]   # mới nhất trước
            changes = [{"d": r.run_date, "src": r.source, "key": r.key, "lb": r.label,
                        "ct": r.change_type, "f": r.field, "old": r.old, "new": r.new}
                       for r in cl.itertuples(index=False)]
        except Exception:
            changes = []
    return {"state": state, "changes": changes}


def tenor_years(period):
    """Kỳ hạn GỐC quy về NĂM (user 17/07/2026: muốn số năm, không phải khoảng).
    Nguồn ghi 3 đơn vị: '3 Năm' | '36 Tháng' | '728 Ngày' -> đều quy về năm."""
    m = re.search(r"(\d+)\s*(Năm|Tháng|Ngày)", str(period or ""), re.I)
    if not m:
        return None
    n, unit = int(m.group(1)), m.group(2).lower()
    y = n if unit == "năm" else (n / 12.0 if unit == "tháng" else n / 365.0)
    return round(y, 1)


def load_dkgd():
    """Danh mục ĐKGD HNX -> {mã CBTT: {tt, gd_dau, gd_cuoi, ma_gd}} (trạng thái + ngày giao dịch).
    Thiếu file -> {} -> cột trạng thái ghi 'Chưa ĐKGD', pipeline vẫn chạy."""
    if not os.path.exists(CATRAW):
        return {}
    c = pd.read_csv(CATRAW, dtype=str).fillna("")
    out = {}
    for _, r in c.iterrows():
        k = (r.get("ma_cbtt") or "").strip().upper()
        if k and k not in out:
            out[k] = {"tt": r.get("trang_thai", "").strip(), "gd_dau": r.get("ngay_gd_dau", "").strip(),
                      "gd_cuoi": r.get("ngay_gd_cuoi", "").strip(), "ma_gd": r.get("ma_gd", "").strip()}
    return out


HNXLIST = "bond_list_raw.csv"


def load_hnx_list():
    """DANH SÁCH TRÁI PHIẾU của HNX (cấp mã) -> DataFrame đã chuẩn hoá số.
    Nguồn: /to-chuc-phat-hanh/danh-sach-trai-phieu (bond_list_scraper.py).

    USER CHỐT 22/07/2026: đây là NGUỒN DƯ NỢ CHÍNH THỨC, thay cho cách tự tính
    (phát hành − mua lại − đáo hạn). Lý do từ đợt đối chiếu 6.795 mã:
      · 4.248/4.702 mã có ở cả hai nguồn KHỚP CHÍNH XÁC từng đồng (1.129,0 nghìn tỷ)
        -> cách tự tính đúng, đổi nguồn không phải vì sai mà vì ĐỘ PHỦ;
      · HNX ghi NGÀY ĐÁO HẠN SAU GIA HẠN (445 mã ta ép về 0 oan = 26,8 nghìn tỷ;
        mức dời 637-731 ngày, không mã nào vượt trần 2 năm của NĐ08/2023);
      · HNX có cả TP phát hành TRƯỚC 2021 (322 mã còn dư nợ = 42,5 nghìn tỷ) mà trang
        "Thông tin phát hành" không lùi tới;
      · 0/2.041 mã còn KL lưu hành bị quá đáo hạn -> nguồn tự làm sạch, KHÁC VSD
        (VSD từng có 101 mã "Hiệu lực" dù quá hạn = 63,6 nghìn tỷ).
    Thiếu file -> trả None -> pipeline lùi về cách tự tính như trước (không gãy).
    """
    if not os.path.exists(HNXLIST):
        return None
    h = pd.read_csv(HNXLIST)
    h["ma_tp"] = h["ma_tp"].astype(str).str.strip()
    h["_k"] = h["ma_tp"].str.upper()
    h = h.drop_duplicates("_k")
    mg = h["menh_gia_num"].fillna(0)
    h["hnx_face"] = h["kl_phat_hanh_num"].fillna(0) * mg
    h["hnx_rem"] = h["kl_con_luu_hanh_num"].fillna(0) * mg
    h["hnx_dh"] = pd.to_datetime(h["ngay_dao_han"], format="%d/%m/%Y", errors="coerce")
    h["hnx_ph"] = pd.to_datetime(h["ngay_phat_hanh"], format="%d/%m/%Y", errors="coerce")
    return h


def compute_outstanding(df, bb):
    """Dư nợ đang lưu hành cấp mã TP.
       Dư nợ LẤY TỪ 'KL còn lưu hành' của danh sách trái phiếu HNX (xem load_hnx_list);
       cách tự tính (đáo hạn->0 / còn lại sau mua lại / giá trị phát hành) được GIỮ LẠI ở
       cột `du_tinh` làm LỚP ĐỐI CHIẾU, và là phương án dự phòng cho mã HNX không liệt kê."""
    today = pd.Timestamp(datetime.now().date())
    iss = df.copy()
    iss["dh"] = pd.to_datetime(iss["ngay_dao_han"], format="%d/%m/%Y", errors="coerce")

    def wcoupon(g):
        v = g["gia_tri_phat_hanh"]; r = g["lai_suat_num"]; mk = r.notna()
        tot = v[mk].sum()
        return (v[mk] * r[mk]).sum() / tot if tot > 0 else None

    face = iss.groupby("ma_tp").agg(
        face=("gia_tri_phat_hanh", "sum"), dh=("dh", "max"),
        nhom=("nhom", "first"), dn=("ten_dn", "first"),
        khn=("ky_han_nhom", "first"), kyhan=("ky_han", "first"),
        ph=("ph_date", "min")).reset_index()
    coup = (iss.groupby("ma_tp").apply(wcoupon, include_groups=False)
              .rename("coupon").reset_index())
    face = face.merge(coup, on="ma_tp", how="left")
    face["nguon"] = "Phát hành"

    # Bản công bố MUA LẠI mới nhất theo mã (kèm ngày đáo hạn & giá trị phát hành của bảng mua lại)
    bb2 = bb.copy()
    bb2["dh_bb"] = pd.to_datetime(bb2["ngay_dao_han"], format="%d/%m/%Y", errors="coerce")
    bb_latest = (bb2.dropna(subset=["dtm"]).sort_values("dtm")
                    .drop_duplicates("ma_tp", keep="last"))

    # Ghép dư nợ còn lại + ngày đáo hạn (bản mua lại) vào universe phát hành.
    # dh = đáo hạn MUỘN HƠN giữa 2 nguồn: bản mua lại công bố sau thường phản ánh gia hạn/điều
    # chỉnh, cứu các mã bị ngày đáo hạn cũ ở bảng phát hành làm ép remaining=0 oan.
    face = face.merge(bb_latest[["ma_tp", "gt_con_lai_num", "dh_bb"]], on="ma_tp", how="left")
    face["dh"] = face[["dh", "dh_bb"]].max(axis=1)

    # Mã CHỈ có ở bảng mua lại (phát hành trước phạm vi dữ liệu chào bán) -> dựng từ chính bảng mua lại.
    only = bb_latest[~bb_latest["ma_tp"].isin(set(face["ma_tp"]))].copy()
    only["khn"] = only["ky_han"].apply(lambda k: tenor_bucket(k, None))
    only = only.rename(columns={"gt_phat_hanh_num": "face", "dh_bb": "dh", "ten_dn": "dn"})
    only["coupon"] = None
    only["nguon"] = "CBTT mua lại"
    only["kyhan"] = only["ky_han"]
    only["ph"] = pd.to_datetime(only["ngay_phat_hanh"], format="%d/%m/%Y", errors="coerce")
    only = only[["ma_tp", "face", "dh", "nhom", "dn", "khn", "kyhan", "ph", "coupon",
                 "gt_con_lai_num", "nguon"]]

    m = pd.concat([face, only], ignore_index=True)

    # ---- BỔ SUNG UNIVERSE TỪ DANH SÁCH TRÁI PHIẾU HNX (user chốt 22/07/2026) ----------
    # Mã HNX liệt kê mà ta chưa có (chủ yếu phát hành TRƯỚC 2021 — ngoài tầm với của trang
    # "Thông tin phát hành"). Nhóm này KHÔNG có dữ liệu đợt phát hành/mua lại nên chỉ điền
    # được các trường có trong danh sách; `nguon` ghi rõ để bảng chi tiết không gây hiểu nhầm.
    hl = load_hnx_list()
    if hl is not None:
        have = set(m["ma_tp"].astype(str).str.strip().str.upper())
        add = hl[~hl["_k"].isin(have)].copy()
        if len(add):
            add["nhom"] = add["ten_dn"].apply(lambda s: classify_with_source(s)[0])
            add["khn"] = add["ky_han"].apply(lambda k: tenor_bucket(k, None))
            add = add.rename(columns={"hnx_face": "face", "hnx_dh": "dh", "ten_dn": "dn",
                                      "ky_han": "kyhan", "hnx_ph": "ph",
                                      "lai_suat_num": "coupon"})
            add["gt_con_lai_num"] = add["hnx_rem"]
            add["nguon"] = "Danh sách HNX"
            m = pd.concat([m, add[["ma_tp", "face", "dh", "nhom", "dn", "khn", "kyhan",
                                   "ph", "coupon", "gt_con_lai_num", "nguon"]]],
                          ignore_index=True)
            print(f"Danh sách HNX: bổ sung {len(add):,} mã ngoài phạm vi CBTT phát hành "
                  f"({int((add['hnx_rem'] > 0).sum()):,} mã còn dư nợ)")
    # bảng tra theo mã (dùng lại ở phần gia hạn & dư nợ bên dưới)
    _hk = m["ma_tp"].astype(str).str.strip().str.upper()
    if hl is not None:
        hidx = hl.set_index("_k")
        m["hnx_rem"] = _hk.map(hidx["hnx_rem"])
        m["hnx_dh"] = _hk.map(hidx["hnx_dh"])
        m["hnx_face"] = _hk.map(hidx["hnx_face"])
    else:
        m["hnx_rem"] = pd.NA
        m["hnx_dh"] = pd.NaT
        m["hnx_face"] = pd.NA
    m["co_hnx"] = m["hnx_rem"].notna()

    # ---- GIA HẠN (user chốt 20/07/2026 — THAY quyết định 17/07 "giữ ngày gốc").
    # Lý do đảo: quyết định 17/07 dựa trên nhận định "dời ngày theo NGUỒN NGOÀI HNX". Rà lại thấy
    # KHÔNG đúng — bằng chứng gia hạn nằm ngay trong dữ liệu HNX (`bond_latepay_raw`,
    # loai_su_kien='gia_han'); VSD chỉ ghi lại lịch mới. `gh_codes` đã lọc đúng nhóm có CBTT HNX.
    #
    # Ca lộ ra vấn đề — cụm Sovico, cùng một đợt gia hạn 2 năm, CBTT cách nhau 1 ngày:
    #   SVBCH2124001-005 (27/09/2023): HNX ĐÃ cập nhật kỳ hạn -> đáo hạn 2026 -> vào dư nợ.
    #   SVACH2124001-006 (28/09/2023): HNX CHƯA cập nhật -> vẫn ghi 2024 -> remaining=0 -> RỚT
    #   khỏi dư nợ, dù VSD lẫn chính CBTT của HNX đều xác nhận đáo hạn 2026.
    # Cùng một chính sách mà hai nửa của một cụm bị xử lý ngược nhau => phải sửa.
    #
    # `dh`     = ngày đáo hạn GỐC. GIỮ NGUYÊN — `load_giahan()` cần nó để tính HẠN CHÓT PHÁP LÝ
    #            (gốc + 730 ngày) và số ngày được gia hạn. Ghi đè `dh` sẽ phá cả hai.
    # `dh_gh`  = lịch gia hạn mới nhất (chỉ nhận khi HNX có CBTT gia hạn xác nhận).
    # `dh_eff` = ngày đáo hạn HIỆU LỰC = dh_gh nếu có, không thì dh. Dùng cho VÒNG ĐỜI
    #            (remaining, trạng thái, kỳ hạn còn lại, năm đáo hạn) và để hiển thị.
    #
    # 22/07/2026 — NGUỒN NGÀY GIA HẠN ĐỔI: ưu tiên NGÀY ĐÁO HẠN CỦA DANH SÁCH HNX, VSD chỉ
    # còn là dự phòng. Danh sách HNX cập nhật ngày sau gia hạn cho 445 mã (so với 13 mã lấy
    # được từ VSD = phủ gấp 34 lần) và là nguồn CƠ SỞ của dự án chứ không phải nguồn ngoài
    # -> hết lý do phải chờ CBTT xác nhận mới dám dời ngày.
    gh_codes, xr = load_giahan_codes(), load_xref()
    dh_gh = []
    for ma, dh, hdh in zip(m["ma_tp"], m["dh"], m["hnx_dh"]):
        k = str(ma).strip().upper()
        if pd.notna(hdh) and (pd.isna(dh) or hdh > dh):
            dh_gh.append(hdh)                      # HNX tự công bố -> tin ngay
            continue
        vdh = pd.to_datetime((xr.get(k) or {}).get("dh", ""), format="%d/%m/%Y", errors="coerce")
        dh_gh.append(vdh if (k in gh_codes and pd.notna(vdh) and (pd.isna(dh) or vdh > dh)) else pd.NaT)
    m["dh_gh"] = pd.Series(dh_gh, index=m.index)
    m["dh_eff"] = m["dh_gh"].where(m["dh_gh"].notna(), m["dh"])
    n_gh = int(m["dh_gh"].notna().sum())
    if n_gh:
        n_song = int(((m["dh"] < today) & (m["dh_eff"] >= today)).sum())
        print(f"Gia hạn: {n_gh} mã có ngày đáo hạn mới (HNX xác nhận); "
              f"{n_song} mã hết bị coi là đã đáo hạn oan -> tính lại vào dư nợ")

    # ---- DƯ NỢ ----------------------------------------------------------------------
    # `du_tinh` = cách TỰ TÍNH cũ (giữ nguyên logic) -> nay là lớp ĐỐI CHIẾU.
    # Vòng đời tính theo `dh_eff` (đã tính gia hạn), KHÔNG theo `dh` gốc.
    m["du_tinh"] = m["gt_con_lai_num"].where(m["gt_con_lai_num"].notna(), m["face"])
    m.loc[m["dh_eff"] < today, "du_tinh"] = 0
    m["du_tinh"] = m["du_tinh"].clip(lower=0)
    # `remaining` = SỐ CHÍNH THỨC: ưu tiên KL còn lưu hành của HNX; mã HNX không liệt kê
    # (64 mã, chủ yếu vừa phát hành chưa kịp lên danh sách) thì dùng số tự tính.
    m["remaining"] = m["hnx_rem"].where(m["co_hnx"], m["du_tinh"]).astype(float).clip(lower=0)
    # trạng thái đối chiếu HNX ↔ tự tính, dùng cho cột "Nguồn" ở bảng chi tiết
    _tol = 0.1 * TY
    m["khop_hnx"] = [
        ("Không có ở HNX" if not ch else
         ("Khớp" if abs((dt or 0) - (hr or 0)) <= _tol else "Lệch"))
        for ch, dt, hr in zip(m["co_hnx"], m["du_tinh"], m["hnx_rem"])]
    m.loc[m["nguon"] == "Danh sách HNX", "khop_hnx"] = "Chỉ có ở HNX"
    m["klcl_nam"] = ((m["dh_eff"] - today).dt.days / 365).round(1)  # kỳ hạn còn lại (năm)

    def status(r):
        if r["remaining"] > 0:
            return "Đang lưu hành"
        return "Đã đáo hạn" if pd.notna(r["dh_eff"]) and r["dh_eff"] < today else "Đã mua lại hết"
    m["trang_thai"] = m.apply(status, axis=1)
    m["nam_dh"] = m["dh_eff"].dt.year

    # ---- KỲ HẠN GỐC (năm) + NGÀY GIAO DỊCH + TRẠNG THÁI ĐKGD (user 17/07/2026)
    m["kyhan_nam"] = m["kyhan"].apply(tenor_years)
    dk = load_dkgd()
    m["ngay_gd"] = [(dk.get(str(k).upper()) or {}).get("gd_dau", "") for k in m["ma_tp"]]
    _gd_cuoi = [(dk.get(str(k).upper()) or {}).get("gd_cuoi", "") for k in m["ma_tp"]]
    _tt_cat = [(dk.get(str(k).upper()) or {}).get("tt", "") for k in m["ma_tp"]]

    def _tt_dkgd(tt, gd_dau, rem, dh, ml):
        """Trạng thái ĐKGD + LÝ DO hủy (user chốt: 'do mua lại' / 'do đáo hạn').
        Quan sát 17/07: mã hủy ĐKGD mà còn dư nợ đều đáo hạn trong ~1 tuần tới
        -> hủy ĐKGD xảy ra TRƯỚC đáo hạn, đúng mô hình vòng đời => ghi 'do sắp đáo hạn'."""
        if not tt:
            return "Chưa ĐKGD"
        if tt == "Đã hủy ĐKGD":
            if rem <= 0 and pd.notna(dh) and dh < today:
                return "Đã hủy ĐKGD (do đáo hạn)"
            if rem <= 0 and ml > 0:
                return "Đã hủy ĐKGD (do mua lại)"
            if pd.notna(dh) and dh >= today and (dh - today).days <= 45:
                return "Đã hủy ĐKGD (do sắp đáo hạn)"
            if ml > 0:
                return "Đã hủy ĐKGD (do mua lại)"
            return "Đã hủy ĐKGD"
        if tt == "Chờ giao dịch":
            return "Chờ giao dịch"
        return "Đang ĐKGD"

    _ml = bb.groupby("ma_tp")["gt_mua_lai_num"].sum()
    m["tt_dkgd"] = [_tt_dkgd(tt, gd, rem, dh, _ml.get(ma, 0))
                    for tt, gd, rem, dh, ma in
                    zip(_tt_cat, m["ngay_gd"], m["remaining"], m["dh_eff"], m["ma_tp"])]
    m["ngay_gd_cuoi"] = _gd_cuoi

    # ---- LỚP ĐỐI CHIẾU VSD (user chốt 17/07/2026: HNX là CƠ SỞ, VSD chỉ phủ lên)
    # Không đổi bất kỳ số nào của HNX, không thêm mã VSD-only vào universe.
    # Thiếu vsd_bond_raw.csv -> xref rỗng -> mọi mã ghi "Không có ở VSD", pipeline vẫn chạy.
    xref = load_xref()
    dc = [doi_chieu(ma, rem, xref) for ma, rem in zip(m["ma_tp"], m["remaining"])]
    m["khop"] = [x[0] for x in dc]
    m["nguon_vsd"] = [x[1] for x in dc]
    m["gt_vsd"] = [x[2] for x in dc]
    m["chenh_vsd"] = [x[3] for x in dc]
    m["nguon_dc"] = [mo_ta_nguon(k, r / TY, (g or 0))
                     for k, r, g in zip(m["khop"], m["remaining"], m["gt_vsd"])]

    # Universe = toàn bộ mã (phát hành ∪ mua lại) -> aggregate mua lại khớp với tab "Mua lại"
    iss_set = set(m["ma_tp"])
    bb_iss = bb[bb["ma_tp"].isin(iss_set)].copy()
    bb_iss["nam_ml"] = bb_iss["dtm"].dt.year

    def ser(s):
        return [{"k": k, "v": round(v / TY, 1)} for k, v in s.items()]

    active = m[m["remaining"] > 0]
    out = {
        "gross_ty": round(face["face"].sum() / TY, 0),
        "outstanding_ty": round(m["remaining"].sum() / TY, 0),
        "buyback_ty": round(bb_iss["gt_mua_lai_num"].sum() / TY, 0),
        "n_active": int((m["remaining"] > 0).sum()),
        "n_settled": int((m["remaining"] <= 0).sum()),
        "n_buyback_events": int(len(bb_iss)),
        "maturity": ser(active.groupby("nam_dh")["remaining"].sum().sort_index()),
        "by_group": ser(active.groupby("nhom")["remaining"].sum().sort_values(ascending=False)),
        "buyback_year": [{"k": int(k), "v": round(v / TY, 1)}
                         for k, v in bb_iss.dropna(subset=["nam_ml"])
                         .groupby("nam_ml")["gt_mua_lai_num"].sum().sort_index().items()],
        "top_buyback": [{"k": (dn[:30] + "…") if len(dn) > 30 else dn, "full": dn,
                         "v": round(v / TY, 1)}
                        for dn, v in bb_iss.groupby("ten_dn")["gt_mua_lai_num"].sum()
                        .sort_values(ascending=False).head(15).items()],
        "detail": m,
    }
    return out


def build_outstanding_sheets(wb, out):
    m = out["detail"]
    # Sheet: Dư nợ & Mua lại (tổng hợp)
    ws = wb.create_sheet("Dư nợ & Mua lại")
    rows = [
        ["CHỈ TIÊU (universe đợt phát hành)", "GIÁ TRỊ (tỷ VNĐ)"],
        ["Tổng phát hành (gross)", out["gross_ty"]],
        ["Đã mua lại trước hạn (lũy kế)", out["buyback_ty"]],
        ["Dư nợ đang lưu hành (hiện tại) - theo KL còn lưu hành HNX", out["outstanding_ty"]],
        ["Số mã còn lưu hành", out["n_active"]],
        ["Số mã đã tất toán (đáo hạn/mua lại hết)", out["n_settled"]],
        ["Số đợt mua lại (universe phát hành)", out["n_buyback_events"]],
    ]
    for r in rows:
        ws.append(r)
    style_header(ws, 2)
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 22
    for row in range(2, len(rows) + 1):
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).number_format = "#,##0"
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")

    r0 = len(rows) + 3
    ws.cell(row=r0, column=1, value="Dư nợ đang lưu hành theo NĂM ĐÁO HẠN (maturity wall)").font = Font(bold=True)
    ws.cell(row=r0 + 1, column=1, value="Năm đáo hạn"); ws.cell(row=r0 + 1, column=2, value="Dư nợ (tỷ VNĐ)")
    style_header(ws, 2, row=r0 + 1)
    for i, x in enumerate(out["maturity"]):
        ws.cell(row=r0 + 2 + i, column=1, value=x["k"])
        c = ws.cell(row=r0 + 2 + i, column=2, value=x["v"]); c.number_format = "#,##0"
    mw_first = r0 + 2
    ch = BarChart(); ch.type = "col"; ch.title = "Dư nợ theo năm đáo hạn (tỷ VNĐ)"
    data = Reference(ws, min_col=2, min_row=r0 + 1, max_row=r0 + 1 + len(out["maturity"]))
    cats = Reference(ws, min_col=1, min_row=mw_first, max_row=mw_first + len(out["maturity"]) - 1)
    ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
    ch.height = 8; ch.width = 20
    ws.add_chart(ch, "D2")

    # Sheet: Chi tiết dư nợ (cấp mã)
    # Dư nợ để MỘT cột (user 17/07: HNX & VSD trùng nhau thì 2 cột gây rối) — cột "Nguồn" nêu rõ
    # khác biệt khi lệch, còn "Giá trị VSD/Chênh" chỉ điền cho dòng KHÔNG khớp.
    # "Ngày đáo hạn" = ngày HIỆU LỰC (đã tính gia hạn). Cột "Đáo hạn gốc" chỉ điền cho mã ĐÃ GIA HẠN
    # -> đọc là thấy ngay vì sao lệch với bảng phát hành HNX, khỏi phải mở tab Gia hạn.
    d = m[["ma_tp", "dn", "nhom", "face", "remaining", "du_tinh", "kyhan_nam", "ph", "ngay_gd",
           "dh_eff", "dh_gh", "dh", "tt_dkgd", "trang_thai", "nguon_dc", "gt_vsd", "chenh_vsd",
           "nguon"]].copy()
    d["du_tinh"] = (d["du_tinh"] / TY).round(1)
    d.loc[m["khop_hnx"].values == "Khớp", "du_tinh"] = None   # khớp -> để trống
    d["face"] = (d["face"] / TY).round(1)
    d["remaining"] = (d["remaining"] / TY).round(1)
    d["ph"] = d["ph"].dt.strftime("%d/%m/%Y")
    d["dh_eff"] = d["dh_eff"].dt.strftime("%d/%m/%Y")
    d["dh"] = d["dh"].dt.strftime("%d/%m/%Y")
    d.loc[d["dh_gh"].isna(), "dh"] = None      # chưa gia hạn -> để trống, khỏi lặp lại cột bên cạnh
    d = d.drop(columns=["dh_gh"])
    khop_mask = m["khop"].values == "Khớp"
    d.loc[khop_mask, ["gt_vsd", "chenh_vsd"]] = None   # khớp -> để trống, khỏi lặp lại số dư nợ
    d = d.sort_values("remaining", ascending=False)
    d.columns = ["Mã TP", "Tổ chức phát hành", "Nhóm", "Giá trị phát hành (tỷ)",
                 "Dư nợ (tỷ) - HNX công bố", "Dư nợ tự tính (tỷ) - chỉ khi lệch",
                 "Kỳ hạn gốc (năm)", "Ngày phát hành", "Ngày giao dịch",
                 "Ngày đáo hạn", "Đáo hạn gốc - chỉ khi đã gia hạn",
                 "Trạng thái ĐKGD", "Trạng thái dư nợ", "Nguồn",
                 "Giá trị VSD (tỷ) - chỉ khi lệch", "Chênh HNX-VSD (tỷ)", "Xuất xứ dòng (HNX)"]
    ws2 = wb.create_sheet("Chi tiết dư nợ")
    write_df(ws2, d, number_cols=["Giá trị phát hành (tỷ)", "Dư nợ (tỷ) - HNX công bố",
                                  "Dư nợ tự tính (tỷ) - chỉ khi lệch",
                                  "Giá trị VSD (tỷ) - chỉ khi lệch", "Chênh HNX-VSD (tỷ)"])


# ---------- JSON cho dashboard ----------
def _lls(v):
    """Loại lãi suất theo CBTT HNX: Cố định / Thả nổi / Kết hợp; rỗng -> 'Không rõ'."""
    s = "" if v is None or (isinstance(v, float) and pd.isna(v)) else str(v).strip()
    return s if s else "Không rõ"


def build_json(df, out, bb=None, sec=None, rating=None, latepay=None, giahan=None):
    def recs(d):
        return json.loads(d.to_json(orient="records", force_ascii=False))

    # thứ tự nhóm ngành CHUẨN (sector_map.GROUP_ORDER) - hợp nhất nhóm xuất hiện ở
    # phát hành + mua lại + giao dịch thứ cấp + XHTN + chậm trả để màu/bộ lọc ổn định & đủ trên MỌI tab.
    present = set(df["nhom"].unique())
    if bb is not None:
        present |= set(bb["nhom"].unique())
    if sec:
        present |= {v[1] for v in sec["meta"].values()}
    if rating:
        present |= {r["nhom"] for r in rating["rows"]}
    if latepay:
        present |= {r["nhom"] for r in latepay["rows"]}
    if giahan:
        present |= {r["nhom"] for r in giahan["rows"]}
    group_order = order_groups(present)

    # dữ liệu cấp-đợt cho dashboard tự lọc/tổng hợp (năm/quý/tháng theo ngày phát hành)
    r = df.copy()
    r["y"] = r["ph_date"].dt.year
    r["m"] = r["ph_date"].dt.month
    r["q"] = r["ph_date"].dt.quarter
    rows = []
    for _, x in r.iterrows():
        ls = x["lai_suat_num"]
        rows.append({
            "ph": x["ngay_phat_hanh"], "dang": x["ngay_dang_tin"],
            "dh": ("" if pd.isna(x.get("ngay_dao_han")) else x.get("ngay_dao_han", "")),
            "y": int(x["y"]), "m": int(x["m"]), "q": int(x["q"]),
            "dn": x["ten_dn"], "ma": x["ma_tp"], "nhom": x["nhom"], "nghn": x.get("nhom_src", ""),
            "kh": x["ky_han"], "khn": x["ky_han_nhom"], "lsn": x["ls_nhom"],
            # loại lãi suất theo CBTT: Cố định / Thả nổi / Kết hợp (thiếu -> "Không rõ")
            "lls": _lls(x.get("loai_lai_suat")),
            "gt": round(x["gia_tri_ty"], 1),
            "kl": (None if pd.isna(x.get("khoi_luong_num")) else int(x["khoi_luong_num"])),
            "ls": (None if pd.isna(ls) else round(float(ls), 2)),
            "tt": x["tinh_trang"],
        })

    out_json = {k: v for k, v in out.items() if k != "detail"} if out else None

    # dòng cấp-mã cho tab "Đang lưu hành" (chỉ trái phiếu còn dư nợ)
    out_rows = []
    if out is not None:
        # loại lãi suất lấy theo bản phát hành MỚI NHẤT của mã; mã chỉ dựng từ CBTT mua lại
        # (phát hành trước 2021) không có bản ghi phát hành -> "Không rõ".
        _l = df.sort_values("ph_date")
        lls_map = dict(zip(_l["ma_tp"], _l["loai_lai_suat"]))
        od = out["detail"]
        act = od[od["remaining"] > 0].copy()
        for _, x in act.iterrows():
            out_rows.append({
                "ma": x["ma_tp"], "dn": x["dn"], "nhom": x["nhom"],
                "gt": round(x["remaining"] / TY, 1), "face": round(x["face"] / TY, 1),
                "ls": (None if pd.isna(x["coupon"]) else round(float(x["coupon"]), 2)),
                "lls": _lls(lls_map.get(x["ma_tp"])),
                # ngày đáo hạn HIỆU LỰC (đã tính gia hạn); `dhg` chỉ có khi mã đã được gia hạn
                "khn": x["khn"],
                "dh": (x["dh_eff"].strftime("%d/%m/%Y") if pd.notna(x["dh_eff"]) else ""),
                "y": (int(x["dh_eff"].year) if pd.notna(x["dh_eff"]) else 0),
                # quý/tháng ĐÁO HẠN — để tab Lưu hành dùng chung bộ lọc năm/quý/tháng
                "q": (int(x["dh_eff"].quarter) if pd.notna(x["dh_eff"]) else 0),
                "m": (int(x["dh_eff"].month) if pd.notna(x["dh_eff"]) else 0),
                "dhg": (x["dh"].strftime("%d/%m/%Y") if pd.notna(x["dh_gh"]) and pd.notna(x["dh"]) else ""),
                "klcl": (None if pd.isna(x["klcl_nam"]) else float(x["klcl_nam"])),
                "nguon": x["nguon"],
                "kyn": (None if pd.isna(x["kyhan_nam"]) else float(x["kyhan_nam"])),
                "ph": (x["ph"].strftime("%d/%m/%Y") if pd.notna(x["ph"]) else ""),
                "ngd": x["ngay_gd"], "ttdk": x["tt_dkgd"],
                # pandas biến None -> NaN khi cột lẫn số => phải dùng pd.isna, `is None` KHÔNG bắt được
                # (NaN lọt sang JS thành NaN, cột hiện "0" thay vì "–").
                "khop": x["khop"], "nvsd": x["nguon_vsd"],
                # đối chiếu DƯ NỢ CHÍNH THỨC (HNX công bố) với cách tự tính của dự án
                "khnx": x["khop_hnx"],
                "dtin": (None if pd.isna(x["du_tinh"]) else round(float(x["du_tinh"]) / TY, 1)),
                "gvsd": (None if pd.isna(x["gt_vsd"]) else float(x["gt_vsd"])),
                "chvsd": (None if pd.isna(x["chenh_vsd"]) else float(x["chenh_vsd"])),
            })

    # dòng cấp-đợt cho tab "Mua lại" (toàn bộ đợt mua lại công bố)
    bb_rows = []
    if bb is not None:
        b = bb.copy()
        b["y"] = b["dtm"].dt.year
        b["m"] = b["dtm"].dt.month
        b["q"] = b["dtm"].dt.quarter
        for _, x in b.iterrows():
            bb_rows.append({
                "ml": x.get("ngay_mua_lai", ""), "dang": x.get("ngay_dang_tin", ""),
                "dn": x["ten_dn"], "ma": x["ma_tp"], "nhom": x["nhom"],
                "y": (int(x["y"]) if pd.notna(x["y"]) else 0),
                "m": (int(x["m"]) if pd.notna(x["m"]) else 0),
                "q": (int(x["q"]) if pd.notna(x["q"]) else 0),
                "gt": round((x.get("gt_mua_lai_num") or 0) / TY, 1),
                "kl": (None if pd.isna(x.get("sl_mua_lai_num")) else int(x["sl_mua_lai_num"])),
                "conlai": round((x.get("gt_con_lai_num") or 0) / TY, 1),
                "face": round((x.get("gt_phat_hanh_num") or 0) / TY, 1),
                "tt": x.get("tinh_trang", ""),
            })

    # ---- MÃ HIỂN THỊ THỐNG NHẤT (user chốt 22/07/2026) ------------------------------
    # HNX tồn tại 2 hệ mã cho cùng một trái phiếu: mã CBTT (bản công bố phát hành, vd
    # VJCH2101) và mã giao dịch/ĐKGD 8 ký tự (vd VJC12101). Trước đây mỗi tab hiển thị
    # theo hệ mã của nguồn nó dùng -> cùng một TP mang 2 tên khác nhau tuỳ tab.
    # Quy ước: CÓ mã giao dịch thì hiển thị mã giao dịch; không có thì giữ mã CBTT.
    #   `ma`  = mã CBTT — GIỮ NGUYÊN vì mọi phép nối giữa các bộ dữ liệu đều theo mã này
    #   `mgd` = mã giao dịch (rỗng nếu chưa ĐKGD)
    #   `mah` = mã hiển thị = mgd or ma  (cột bảng sắp xếp/lọc theo trường này)
    # Chỉ xuất BẢNG TRA mã CBTT -> mã giao dịch; dashboard tự gắn `mgd`/`mah` cho mọi bộ
    # dữ liệu (kể cả các bảng được dựng bằng JS như danh sách chậm trả) -> một nguồn sự thật.
    gd_map = {k: v["ma_gd"] for k, v in load_dkgd().items() if v.get("ma_gd")}
    print(f"Mã giao dịch (ĐKGD): {len(gd_map)} mã CBTT có mã GD 8 ký tự")

    data = {
        "mgd_map": gd_map,
        "updated": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "data_date": f"{df['post_date'].max():%d/%m/%Y}",
        "groups": group_order,
        "rows": rows,
        "out": out_json,
        "out_rows": out_rows,
        "bb_rows": bb_rows,
        "sec": sec,
        "rating": rating,
        "latepay": latepay,
        "giahan": giahan,
        "updates": load_updates(),
    }
    # NaN/Infinity không phải JSON hợp lệ. Python vẫn ghi ra chữ 'NaN', và vì data được
    # nhúng thành `const DATA={...}` nên trình duyệt đọc thành số NaN -> hỏng thang đo chart.
    # Quét sạch trước khi ghi, rồi dump với allow_nan=False để lần sau lỗi nổ ngay lúc build.
    data = sanitize_nan(data)
    with open("dashboard_data.json", "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, allow_nan=False)
    print("Đã tạo dashboard_data.json")

    # nhúng data vào dashboard.html (self-contained, mở offline được)
    try:
        with open("dashboard_template.html", encoding="utf-8") as f:
            tpl = f.read()
        html = tpl.replace("__DATA__", json.dumps(data, ensure_ascii=False, allow_nan=False))
        with open("dashboard.html", "w", encoding="utf-8") as f:
            f.write(html)
        print("Đã tạo dashboard.html")
    except FileNotFoundError:
        print("(!) Không thấy dashboard_template.html - bỏ qua dashboard.html")
    return data


if __name__ == "__main__":
    df = load()
    out = None
    if os.path.exists(BBRAW):
        bb = load_buyback()
        out = compute_outstanding(df, bb)
        print(f"Dư nợ đang lưu hành: {out['outstanding_ty']:,.0f} tỷ "
              f"| Gross: {out['gross_ty']:,.0f} tỷ | Mua lại: {out['buyback_ty']:,.0f} tỷ")
    else:
        print("(!) Chưa có bond_buyback_raw.csv - bỏ qua phần dư nợ. Chạy bond_buyback_scraper.py trước.")
    sec = load_secondary(df, bb if out is not None else None)
    if sec is None:
        print("(!) Chưa có bond_secondary_raw.csv - bỏ qua tab Giao dịch thứ cấp.")
    rating = load_rating()
    if rating is None:
        print("(!) Chưa có bond_rating_raw.csv - bỏ qua tab Xếp hạng tín nhiệm.")
    dno_map = outstanding_by_bond(df, bb if out is not None else None)
    news = load_news()
    latepay = load_latepay(dno_map, out["outstanding_ty"] if out is not None else None, news)
    if latepay is None:
        print("(!) Chưa có bond_latepay_raw.csv - bỏ qua tab Chậm trả gốc/lãi.")
    build_excel(df, out)
    giahan = load_giahan(out)
    build_json(df, out, bb if out is not None else None, sec, rating, latepay, giahan)
    print("Hoàn tất.")
