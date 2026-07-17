# -*- coding: utf-8 -*-
"""Cap nhat XLSX tu CSV da backfill (chay sau backfill_2021_2023.py)."""
import csv, os, re
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
CSV_PATH = os.path.join(HERE, "TPDN_VBMA_tong_hop_thang.csv")
XLSX_PATH = os.path.join(HERE, "TPDN_VBMA_tong_hop_thang.xlsx")

with open(CSV_PATH, encoding="utf-8-sig", newline="") as f:
    rdr = list(csv.reader(f))
all_rows = [r for r in rdr[1:] if r and r[0].strip()]
assert all_rows[0][0] == "2021-01" and len(all_rows) == 66, "CSV chua backfill?"

wb = load_workbook(XLSX_PATH)
ws = wb["TPDN theo thang"]
# go merge trong vung du lieu (giu merge tieu de dong 1-2)
for rng in list(ws.merged_cells.ranges):
    if rng.min_row >= 3:
        ws.unmerge_cells(str(rng))
ws["A1"] = ("TONG HOP TRAI PHIEU DOANH NGHIEP (TPDN) - Nguon: Bao cao thang VBMA (T1/2021 - T6/2026)"
            "  [2021: cong bo lan dau; 2022: dinh chinh M+12 (T1-T2 lan dau); 2023+: dinh chinh M+1]")
# xoa vung cu roi ghi lai
for row in ws.iter_rows(min_row=4, max_row=max(ws.max_row, 4 + len(all_rows)), max_col=21):
    for c in row:
        c.value = None
for i, r in enumerate(all_rows):
    for j, v in enumerate(r):
        x = v.strip()
        if x:
            try:
                x = float(x) if "." in x else int(x)
            except ValueError:
                pass
        ws.cell(row=4 + i, column=1 + j, value=(x if x != "" else None))

wsb = wb["Bieu do"]
for rng in list(wsb.merged_cells.ranges):
    wsb.unmerge_cells(str(rng))
for i, r in enumerate(all_rows):
    wsb.cell(row=2 + i, column=1, value=r[0])
    wsb.cell(row=2 + i, column=2, value=float(r[3]) if r[3].strip() else 0)
    wsb.cell(row=2 + i, column=3, value=float(r[5]) if r[5].strip() else 0)
last = 1 + len(all_rows)
for ch in wsb._charts:
    for ser in ch.series:
        if ser.val and ser.val.numRef:
            ser.val.numRef.f = re.sub(r"\$(\d+)$", f"${last}", ser.val.numRef.f)
        if ser.cat and ser.cat.numRef:
            ser.cat.numRef.f = re.sub(r"\$(\d+)$", f"${last}", ser.cat.numRef.f)
        if ser.cat and getattr(ser.cat, "strRef", None) and ser.cat.strRef:
            ser.cat.strRef.f = re.sub(r"\$(\d+)$", f"${last}", ser.cat.strRef.f)

wsg = wb["Ghi chu & Nguon"]
for row in wsg.iter_rows(min_col=1, max_col=1):
    c = row[0]
    if isinstance(c.value, str) and c.value.startswith("Pham vi:"):
        c.value = "Pham vi: 66 thang, T1/2021 - T6/2026 (67 bao cao thang VBMA; 36 thang 2021-2023 bo sung 16/07/2026)."
last_row = wsg.max_row
notes = [
    "",
    "PHUONG PHAP THEO THOI KY (bo sung 16/07/2026):",
    "- 2021: bao cao VBMA chua co Phu luc 1 TPDN -> so CONG BO LAN DAU (trang TPDN bao cao thang do).",
    "- 2022: so DINH CHINH lay tu cot 'cung ky nam truoc' Phu luc 1 bao cao 2023 (M+12).",
    "  Rieng T1-T2/2022 khong co so dinh chinh -> cong bo lan dau. T5/2022 dung chenh luy ke.",
    "- 2023 tro di: so DINH CHINH chuan M+1 (cot 'thang truoc' Phu luc 1 bao cao ke tiep).",
    "- Mua lai: co so thang tu T1/2023 (2022 chi co luy ke nam = 210,830 ty). Cham tra: tu T10/2023.",
    "- GTGD thu cap TPDN rieng le: VBMA dua vao bao cao tu T3/2024 -> 2021-2023 de trong.",
    "- Moc kiem chung (PL1 bao cao T12/2023): 2022 RL 249,075/CC 21,237; 2023 RL 274,170/CC 37,070.",
    "  Tong 12 thang cua chuoi thap hon moc nam do CBTT muon (dac diem du lieu, khong phai loi).",
    "- %YoY chi tinh tu 2023-03 (ca 2 nam deu la so dinh chinh); %MoM bo qua khi thang truoc <500 ty.",
]
for i, s in enumerate(notes):
    wsg.cell(row=last_row + 1 + i, column=1, value=(s or None))
wb.save(XLSX_PATH)
print("XLSX OK:", len(all_rows), "dong")
