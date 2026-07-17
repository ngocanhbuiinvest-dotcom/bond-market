# -*- coding: utf-8 -*-
"""
Backfill TPDN_VBMA_tong_hop_thang.{csv,xlsx} voi 36 thang T1/2021 - T12/2023.
Nguon: 37 bao cao thang VBMA (T1/2021 - T1/2024), cao 16/07/2026 qua in-app browser + pdf.js.

PHUONG PHAP THEO THOI KY (do format bao cao VBMA thay doi):
- 2021:      bao cao chua co Phu luc 1 TPDN -> dung SO CONG BO LAN DAU tu trang 'Trai phieu
             doanh nghiep' cua bao cao thang do. (T10/2021: CC = chenh luy ke 10T-8T = 3,966;
             RL = tong 39,285 - CC.)
- 2022:      Phu luc 1 TPDN chi xuat hien tu bao cao T3/2023 -> lay SO DINH CHINH tu cot
             "cung ky nam truoc" (M+12) cua bao cao 2023. Rieng T5/2022 cot thang bi lap so
             (29,262) -> dung chenh luy ke 5T-4T/2022 = 39,977 (RL), 300 (CC).
             T1-T2/2022: khong co so dinh chinh -> so cong bo lan dau (luy ke dinh chinh
             T1+T2/2022 = RL 35,765 / CC 7,205 nhung khong co phan bo thang).
- 2023:      SO DINH CHINH chuan M+1: cot "thang truoc" Phu luc 1 bao cao ke tiep
             (T1/2023 lay tu cot cung ky bao cao T1/2024).
- Mua lai:   VBMA chi cong bo so thang tu T1/2023. 2021-2022 de trong
             (nam 2022 luy ke cong bo = 210,830 ty).
- Cham tra:  VBMA cong bo so ma/gia tri theo thang tu ~T10/2023. Truoc do de trong.
- GTGD thu cap TPDN rieng le: VBMA chua dua vao bao cao giai doan nay -> trong (nhu T1-T2/2024).

MOC KIEM CHUNG (Phu luc 1 bao cao T12/2023): nam 2022 = RL 249,075 / CC 21,237 / Tong 270,312;
nam 2023 = RL 274,170 / CC 37,070 / Tong 311,240. Tong thang cua chuoi se THAP hon moc nam
(2022 thieu dinh chinh T1-T2; 2023 dinh chinh M+1 chua bat het CBTT muon) - la dac diem du lieu.
"""
import csv, shutil, datetime, calendar, os

HERE = os.path.dirname(os.path.abspath(__file__))
CSV_PATH = os.path.join(HERE, "TPDN_VBMA_tong_hop_thang.csv")
XLSX_PATH = os.path.join(HERE, "TPDN_VBMA_tong_hop_thang.xlsx")
SUF = "_backup_20260716_truoc_backfill2021"

# ---- Du lieu 2021-2023: (thang, RL, so_dot_RL, CC, so_dot_CC, mua_lai, so_ma_cham, goc_cham) ----
# None = khong co du lieu trong bao cao VBMA
ROWS = [
    # 2021 - cong bo lan dau (trang TPDN bao cao thang do)
    ("2021-01",  6416,   7,  1610,    2, None, None, None),
    ("2021-02",  1040,   3,  1515,    1, None, None, None),
    ("2021-03",  5175,  17,  2860,    2, None, None, None),
    ("2021-04",  29579.6, 36, 0,      0, None, None, None),
    ("2021-05",  28410, 46,  500,     1, None, None, None),
    ("2021-06",  52274, 91,  1500,    1, None, None, None),
    ("2021-07",  38905, 53,  0,       0, None, None, None),
    ("2021-08",  24077, 51,  2000,    1, None, None, None),
    ("2021-09",  29734, 42,  0,       0, None, None, None),
    ("2021-10",  35319, 53,  3966,    4, None, None, None),  # CC = luy ke 10T(15,550)-8T(11,584); RL = 39,285-CC
    ("2021-11",  18276, 40,  2090,    1, None, None, None),
    ("2021-12",  65757, 80,  0,       0, None, None, None),
    # 2022 - dinh chinh theo cot cung ky (M+12) Phu luc 1 bao cao 2023; so dot = cong bo lan dau
    ("2022-01",  19500, 16,  5009,    7, None, None, None),  # cong bo lan dau (khong co so dinh chinh)
    ("2022-02",  1300,   4,  500,     1, None, None, None),  # cong bo lan dau
    ("2022-03",  26844, 10,  1491,    2, None, None, None),  # PL1 bao cao T3/2023
    ("2022-04",  29262, 23,  0,       0, None, None, None),  # PL1 T4/2023 (khop chenh luy ke)
    ("2022-05",  39977, 34,  300,     1, None, None, None),  # chenh luy ke 5T-4T/2022 (cot thang bi lap 29,262)
    ("2022-06",  43593, 44,  0,       0, None, None, None),  # PL1 T6/2023
    ("2022-07",  26195, 28,  0,       0, None, None, None),  # PL1 T7/2023
    ("2022-08",  21419, 26,  1268,    1, None, None, None),  # PL1 T8/2023
    ("2022-09",  17540, 25,  335,     1, None, None, None),  # PL1 T9/2023 (CC = Tong-RL)
    ("2022-10",  335,    1,  0,       0, None, None, None),  # PL1 T10/2023
    ("2022-11",  2077,   5,  0,       0, None, None, None),  # PL1 T11/2023
    ("2022-12",  10638, 11,  10638, None, None, None, None), # PL1 T12/2023 (CC 10,637.8 BIDV cong bo T1/2023)
    # 2023 - dinh chinh M+1 (cot "thang truoc" PL1 bao cao ke tiep); mua lai tu trang TPDN
    ("2023-01",  110,  None, 21,   None, 8068,  None, None), # cot cung ky PL1 bao cao T1/2024
    ("2023-02",  500,   1,   1500,   2,  4782,  None, None), # PL1 T3/2023
    ("2023-03",  24425, 10,  2000, None, 14267, None, None), # PL1 T4/2023
    ("2023-04",  671,   1,   2000,   1,  11398, None, None), # PL1 T5/2023
    ("2023-05",  2600,  4,   0,      0,  22789, None, None), # PL1 T6/2023
    ("2023-06",  8170,  13,  0,      0,  31591, None, None), # PL1 T7/2023
    ("2023-07",  33135, 14,  7500,   7,  20533, None, None), # PL1 T8/2023
    ("2023-08",  32791, 22,  1813, None, 17489, None, None), # PL1 T9/2023 (CC = Tong-RL)
    ("2023-09",  37509, 14,  3344, None, 9249,  None, None), # PL1 T10/2023 (CC = Tong-RL)
    ("2023-10",  28056, 18,  3092, None, 13645, 16, 1006),   # PL1 T11/2023; cham tra: trang TPDN T10/2023
    ("2023-11",  35965, 27,  10000,None, 8754,  None, None), # PL1 T12/2023 (CC = Tong-RL)
    ("2023-12",  66980, 55,  2813, None, 32677, 6, 545.7),   # PL1 T1/2024; cham tra: trang TPDN T12/2023
]

def eom(t):
    y, m = int(t[:4]), int(t[5:7])
    return f"{calendar.monthrange(y,m)[1]:02d}/{m:02d}/{y}"

def fnum(v):
    if v is None: return ""
    if isinstance(v, float) and v == int(v): return str(int(v))
    return str(v)

# ---- Doc CSV hien tai ----
with open(CSV_PATH, encoding="utf-8-sig", newline="") as f:
    rdr = list(csv.reader(f))
header, old = rdr[0], [r for r in rdr[1:] if r and r[0].strip()]
assert old[0][0] == "2024-01", "CSV khong bat dau tu 2024-01: " + old[0][0]

# ---- Backup ----
shutil.copy2(CSV_PATH, CSV_PATH.replace(".csv", SUF + ".csv"))
shutil.copy2(XLSX_PATH, XLSX_PATH.replace(".xlsx", SUF + ".xlsx"))
print("Da backup voi hau to", SUF)

# ---- Dung 36 dong moi ----
new_rows, lkRL, lkCC, lkT, prev_tong, year = [], 0, 0, 0, None, None
tong_by_month = {}
for (t, rl, sdrl, cc, sdcc, ml, scham, gcham) in ROWS:
    y = t[:4]
    if y != year: year, lkRL, lkCC, lkT = y, 0, 0, 0
    rl_, cc_ = rl or 0, cc or 0
    tong = rl_ + cc_
    lkRL += rl_; lkCC += cc_; lkT += tong
    tong_by_month[t] = tong
    mom = ""
    if prev_tong is not None and prev_tong >= 500:
        mom = str(round((tong / prev_tong - 1) * 100))
    # %YoY: chi tinh khi ca 2 nam deu trong vung dinh chinh (2023-03+ vs 2022-03+)
    yoy = ""
    if t >= "2023-03":
        base = tong_by_month.get(f"{int(y)-1}{t[4:]}")
        if base and base >= 500:
            yoy = str(round((tong / base - 1) * 100))
    m = int(t[5:7])
    new_rows.append([
        t, f"T{m}/{y}", eom(t),
        fnum(rl), fnum(sdrl), fnum(cc), fnum(sdcc), fnum(tong),
        mom, yoy, fnum(lkRL), fnum(lkCC), fnum(lkT),
        fnum(ml), fnum(scham), fnum(gcham),
        "", "", "", "", "",   # GTGD/BQ/%MoM/%YoY/luy ke thu cap: VBMA chua cong bo giai doan nay
    ])
    prev_tong = tong

# ---- Cap nhat dong 2024 hien co: %MoM 2024-01 va %YoY 2024 (goc 2023 da co) ----
for r in old:
    t = r[0]
    if t == "2024-01" and not r[8].strip():
        r[8] = str(round((float(r[7]) / prev_tong - 1) * 100))  # prev = T12/2023
    if t[:4] == "2024" and not r[9].strip():
        base = tong_by_month.get("2023" + t[4:])
        if base and base >= 500:
            r[9] = str(round((float(r[7]) / base - 1) * 100))

# ---- Ghi CSV ----
with open(CSV_PATH, "w", encoding="utf-8-sig", newline="") as f:
    w = csv.writer(f)
    w.writerow(header)
    w.writerows(new_rows)
    w.writerows(old)
print(f"CSV: {len(new_rows)} dong moi + {len(old)} dong cu = {len(new_rows)+len(old)} thang")

# ---- Kiem chung moc nam ----
s22rl = sum(r[1] for r in ROWS if r[0][:4] == "2022")
s22cc = sum(r[3] for r in ROWS if r[0][:4] == "2022")
s23rl = sum(r[1] for r in ROWS if r[0][:4] == "2023")
s23cc = sum(r[3] for r in ROWS if r[0][:4] == "2023")
print(f"KIEM CHUNG 2022: RL {s22rl:,.0f} vs moc 249,075 (chenh {s22rl-249075:+,.0f} do T1-T2 chua dinh chinh)")
print(f"                 CC {s22cc:,.0f} vs moc 21,237  (chenh {s22cc-21237:+,.0f})")
print(f"KIEM CHUNG 2023: RL {s23rl:,.0f} vs moc 274,170 (chenh {s23rl-274170:+,.0f} = CBTT muon sau M+1)")
print(f"                 CC {s23cc:,.0f} vs moc 37,070  (chenh {s23cc-37070:+,.0f})")

# ---- Cap nhat XLSX ----
from openpyxl import load_workbook
wb = load_workbook(XLSX_PATH)
ws = wb["TPDN theo thang"]
ws["A1"] = ("TONG HOP TRAI PHIEU DOANH NGHIEP (TPDN) - Nguon: Bao cao thang VBMA (T1/2021 - T6/2026)"
            "  [2021: cong bo lan dau; 2022: dinh chinh M+12 (T1-T2 lan dau); 2023+: dinh chinh M+1]")
# ghi lai toan bo khoi du lieu tu dong 4
all_rows = new_rows + old
for i, r in enumerate(all_rows):
    for j, v in enumerate(r):
        x = v if isinstance(v, (int, float)) else (v.strip() if isinstance(v, str) else v)
        if isinstance(x, str) and x:
            try: x = float(x) if "." in x else int(x)
            except ValueError: pass
        ws.cell(row=4+i, column=1+j, value=(x if x != "" else None))
# sheet Bieu do: bang A:C + cap nhat range chart
wsb = wb["Bieu do"]
for i, r in enumerate(all_rows):
    wsb.cell(row=2+i, column=1, value=r[0])
    wsb.cell(row=2+i, column=2, value=float(r[3]) if r[3] else 0)
    wsb.cell(row=2+i, column=3, value=float(r[5]) if r[5] else 0)
last = 1 + len(all_rows)
import re as _re
for ch in wsb._charts:
    for ser in ch.series:
        if ser.val and ser.val.numRef:
            ser.val.numRef.f = _re.sub(r"\$(\d+)$", f"${last}", ser.val.numRef.f)
        if ser.cat and ser.cat.numRef:
            ser.cat.numRef.f = _re.sub(r"\$(\d+)$", f"${last}", ser.cat.numRef.f)
        if ser.cat and getattr(ser.cat, "strRef", None):
            ser.cat.strRef.f = _re.sub(r"\$(\d+)$", f"${last}", ser.cat.strRef.f)
# Ghi chu: pham vi + phuong phap theo thoi ky
wsg = wb["Ghi chu & Nguon"]
for row in wsg.iter_rows(min_col=1, max_col=1):
    c = row[0]
    if isinstance(c.value, str) and c.value.startswith("Pham vi:"):
        c.value = "Pham vi: 66 thang, T1/2021 - T6/2026 (67 bao cao thang VBMA; 36 thang 2021-2023 bo sung 16/07/2026)."
last_row = wsg.max_row
notes = [
    "",
    "PHUONG PHAP THEO THOI KY (bo sung 16/07/2026):",
    "- 2021: bao cao VBMA chua co Phu luc 1 TPDN -> so CONG BO LAN DAU (trang TPDN bao cao thang do).",
    "- 2022: so DINH CHINH lay tu cot 'cung ky nam truoc' Phu luc 1 bao cao 2023 (M+12).",
    "  Rieng T1-T2/2022 khong co so dinh chinh -> cong bo lan dau. T5/2022 dung chenh luy ke.",
    "- 2023 tro di: so DINH CHINH chuan M+1 (cot 'thang truoc' Phu luc 1 bao cao ke tiep).",
    "- Mua lai: co so thang tu T1/2023 (2022 chi co luy ke nam = 210,830 ty). Cham tra: tu T10/2023.",
    "- GTGD thu cap TPDN rieng le: VBMA dua vao bao cao tu T3/2024 -> 2021-2023 de trong.",
    "- Moc kiem chung (PL1 bao cao T12/2023): 2022 RL 249,075/CC 21,237; 2023 RL 274,170/CC 37,070.",
    "  Tong 12 thang cua chuoi thap hon moc nam do CBTT muon (dac diem du lieu, khong phai loi).",
    "- %YoY chi tinh tu 2023-03 (ca 2 nam deu la so dinh chinh); %MoM bo qua khi thang truoc <500 ty.",
]
for i, s in enumerate(notes):
    wsg.cell(row=last_row + 1 + i, column=1, value=(s or None))
wb.save(XLSX_PATH)
print("XLSX da cap nhat:", len(all_rows), "dong du lieu")
