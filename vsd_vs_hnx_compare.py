# -*- coding: utf-8 -*-
"""
ĐỐI CHIẾU TPDN riêng lẻ: VSD (vsd.vn/vi/ibl) vs HNX (cbonds.hnx.vn)

Hai nguồn đo hai thứ KHÁC NHAU — mục tiêu là GIẢI THÍCH chênh lệch, không phải ép khớp 100%:
  - VSD  = ĐĂNG KÝ LƯU KÝ tại VSDC (Giấy CNĐKCK). Còn "Hiệu lực" / đã "Hủy đăng ký".
  - HNX  = (a) Danh mục ĐĂNG KÝ GIAO DỊCH trên sàn TPRL (bond_catalog_raw)
           (b) CBTT phát hành / mua lại (bond_issuance / bond_buyback, dữ liệu từ ~30/12/2020)
  Một mã có thể đã lưu ký (VSD) nhưng chưa/không ĐKGD (HNX) và ngược lại.

KHÓA NỐI (bài học của dự án: mã GD 'VJC12101' != mã CBTT 'VJCH2101'):
  VSD cho cả hai — mã CK = mã GD, và mã CBTT nằm trong tên chứng khoán.
  Nối: ISIN -> mã CBTT -> mã GD.
  LƯU Ý: với trái phiếu phát hành gần đây, mã CBTT TRÙNG mã GD (vd ABB12501, P5332601)
  => khi đối chiếu sang bảng CBTT phát hành phải thử CẢ HAI khóa (cbtt_k và gd_k),
     nếu không sẽ báo nhầm là "HNX có, VSD không có".

Đầu vào : vsd_bond_raw.csv, bond_catalog_raw.csv, bond_issuance_raw.csv, bond_buyback_raw.csv
Đầu ra  : VSD_vs_HNX_DoiChieu.xlsx (9 sheet) + tóm tắt ra màn hình
"""
import re
import sys
from datetime import date, datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

VSD_CSV = "vsd_bond_raw.csv"
CAT_CSV = "bond_catalog_raw.csv"
ISS_CSV = "bond_issuance_raw.csv"
BB_CSV = "bond_buyback_raw.csv"
OUT_XLSX = "VSD_vs_HNX_DoiChieu.xlsx"

TY = 1_000_000_000
NGHIN_TY = TY * 1000
TODAY = date.today()


def _num(s):
    """HNX '1,174' (phẩy nghìn) | VSD '1.174' (chấm nghìn) -> bỏ mọi ký tự ngăn cách."""
    if pd.isna(s) or s == "":
        return 0
    s = re.sub(r"[^\d]", "", str(s))
    return int(s) if s else 0


def _date(s):
    if pd.isna(s) or not str(s).strip():
        return None
    try:
        return datetime.strptime(str(s).strip(), "%d/%m/%Y").date()
    except ValueError:
        return None


def _dstr(d):
    return d.strftime("%d/%m/%Y") if isinstance(d, date) else ""


def _matured(d):
    """groupby của pandas trả NaN (float) cho ngày thiếu -> phải kiểm tra kiểu."""
    return isinstance(d, date) and d < TODAY


# --------------------------------------------------------------- nạp dữ liệu
def load_vsd():
    df = pd.read_csv(VSD_CSV, dtype=str, encoding="utf-8-sig").fillna("")
    df["sl"] = df["so_luong"].map(_num)
    df["mg"] = df["menh_gia"].map(_num)
    df["gt"] = df["gia_tri_dk"].map(_num)
    miss = (df["gt"] == 0) & (df["sl"] > 0) & (df["mg"] > 0)
    df.loc[miss, "gt"] = df.loc[miss, "sl"] * df.loc[miss, "mg"]
    df["hieu_luc"] = df["tinh_trang"].str.strip().str.lower().eq("hiệu lực")
    df["d_ph"] = df["ngay_phat_hanh"].map(_date)
    df["d_dh"] = df["ngay_dao_han"].map(_date)
    df["isin_k"] = df["isin"].str.strip().str.upper()
    df["cbtt_k"] = df["ma_cbtt"].str.strip().str.upper()
    df["gd_k"] = df["ma_ck"].str.strip().str.upper()
    return df


def load_catalog():
    df = pd.read_csv(CAT_CSV, dtype=str, encoding="utf-8-sig").fillna("")
    df["kl"] = df["kl_dkgd"].map(_num)
    df["mg"] = df["menh_gia"].map(_num)
    df["gt"] = df["kl"] * df["mg"]
    df["dkgd"] = df["trang_thai"].str.strip().str.lower().eq("đăng ký giao dịch")
    df["isin_k"] = df["isin"].str.strip().str.upper()
    df["cbtt_k"] = df["ma_cbtt"].str.strip().str.upper()
    df["gd_k"] = df["ma_gd"].str.strip().str.upper()
    return df


def load_issuance():
    df = pd.read_csv(ISS_CSV, dtype=str, encoding="utf-8-sig").fillna("")
    df["gt"] = df["khoi_luong"].map(_num) * df["menh_gia"].map(_num)
    df["d_ph"] = df["ngay_phat_hanh"].map(_date)
    df["d_dh"] = df["ngay_dao_han"].map(_date)
    df["cbtt_k"] = df["ma_tp"].str.strip().str.upper()
    df["_dt"] = df["ngay_dang_tin"].map(_date)
    # gộp trùng theo (mã TP + ngày phát hành), giữ bản đăng tin mới nhất — quy ước dự án
    df = (df.sort_values("_dt")
            .drop_duplicates(subset=["cbtt_k", "ngay_phat_hanh"], keep="last"))
    return (df.groupby("cbtt_k")
              .agg(gt_ph=("gt", "sum"), ten_dn=("ten_dn", "last"),
                   i_ph=("d_ph", "min"), i_dh=("d_dh", "max"), n_dot=("gt", "size"))
              .reset_index())


def load_buyback():
    df = pd.read_csv(BB_CSV, dtype=str, encoding="utf-8-sig").fillna("")
    df["cbtt_k"] = df["ma_tp"].str.strip().str.upper()
    df["gt_ml"] = df["gt_mua_lai"].map(_num)
    df["_dt"] = df["ngay_dang_tin"].map(_date)
    df = df.sort_values("_dt").drop_duplicates(
        subset=["cbtt_k", "ngay_mua_lai", "gt_mua_lai"], keep="last")
    return df.groupby("cbtt_k")["gt_ml"].sum().reset_index()


# ------------------------------------------------------------------- nối khóa
def join_vsd_hnx(vsd, cat):
    """Nối VSD <-> danh mục ĐKGD HNX theo ISIN, vá bằng mã CBTT, rồi mã GD."""
    cat_i = set(cat["isin_k"])
    cat_c = cat[cat["cbtt_k"] != ""].drop_duplicates("cbtt_k").set_index("cbtt_k")["isin_k"]
    cat_g = cat[cat["gd_k"] != ""].drop_duplicates("gd_k").set_index("gd_k")["isin_k"]
    keys, how = [], []
    for isin, cbtt, gd in zip(vsd["isin_k"], vsd["cbtt_k"], vsd["gd_k"]):
        if isin and isin in cat_i:
            keys.append(isin); how.append("ISIN")
        elif cbtt and cbtt in cat_c.index:
            keys.append(cat_c[cbtt]); how.append("Mã CBTT")
        elif gd and gd in cat_g.index:
            keys.append(cat_g[gd]); how.append("Mã GD")
        else:
            keys.append(""); how.append("")
    vsd = vsd.copy()
    vsd["join_isin"], vsd["join_how"] = keys, how
    return vsd


# -------------------------------------------------------------------- Excel
H_FILL = PatternFill("solid", fgColor="1F4E79")
H_FONT = Font(color="FFFFFF", bold=True)
SEC_FONT = Font(bold=True, color="1F4E79")
MONEY = ("tỷ", "Giá trị", "Số lượng", "Chênh", "Mệnh giá", "KL")


def write_sheet(wb, name, df, note=None, widths=None):
    ws = wb.create_sheet(name[:31])
    r0 = 1
    if note:
        c = ws.cell(1, 1, note)
        c.font = Font(italic=True, color="555555")
        r0 = 3
    for j, col in enumerate(df.columns, 1):
        c = ws.cell(r0, j, str(col))
        c.fill, c.font = H_FILL, H_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    for i, (_, row) in enumerate(df.iterrows(), r0 + 1):
        for j, v in enumerate(row, 1):
            if isinstance(v, float) and pd.isna(v):
                v = ""
            ws.cell(i, j, v)
    if len(df):
        ws.freeze_panes = ws.cell(r0 + 1, 1)
        ws.auto_filter.ref = f"A{r0}:{get_column_letter(len(df.columns))}{r0 + len(df)}"
    for j, col in enumerate(df.columns, 1):
        w = (widths or {}).get(col)
        if w is None:
            vals = [len(str(col))] + [len(str(x)) for x in df[col].head(300)]
            w = min(44, max(10, max(vals) + 2))
        ws.column_dimensions[get_column_letter(j)].width = w
        if any(k in str(col) for k in MONEY):
            fmt = "#,##0.00" if "tỷ" in str(col) or "Chênh giá trị" in str(col) else "#,##0"
            for i in range(r0 + 1, r0 + 1 + len(df)):
                ws.cell(i, j).number_format = fmt
    return ws


def main():
    print("Nạp dữ liệu...")
    vsd, cat, iss, bb = load_vsd(), load_catalog(), load_issuance(), load_buyback()
    print(f"  VSD {len(vsd)} mã | HNX ĐKGD {len(cat)} mã | "
          f"HNX CBTT phát hành {len(iss)} mã | mua lại {len(bb)} mã")

    vsd = join_vsd_hnx(vsd, cat)
    ci = cat.set_index("isin_k")
    matched = vsd[vsd["join_isin"] != ""].copy()
    only_vsd = vsd[vsd["join_isin"] == ""].copy()
    only_hnx = cat[~cat["isin_k"].isin(set(matched["join_isin"]))].copy()

    m = matched.join(ci[["ma_cbtt", "ma_gd", "ten_tcph", "kl", "mg", "gt",
                         "trang_thai", "dkgd"]].add_prefix("h_"), on="join_isin")
    m = m.merge(iss.rename(columns={"cbtt_k": "h_ma_cbtt"})[
        ["h_ma_cbtt", "gt_ph", "n_dot", "i_ph", "i_dh"]], on="h_ma_cbtt", how="left")
    m = m.merge(bb.rename(columns={"cbtt_k": "h_ma_cbtt", "gt_ml": "gt_ml"}),
                on="h_ma_cbtt", how="left")
    m["gt_ph"] = m["gt_ph"].fillna(0)
    m["gt_ml"] = m["gt_ml"].fillna(0)

    # nhóm cầu nối
    A = m[m["hieu_luc"] & m["h_dkgd"]]            # cả 2 còn hiệu lực
    B = m[m["hieu_luc"] & ~m["h_dkgd"]]           # VSD hiệu lực, HNX đã hủy ĐKGD
    C = m[~m["hieu_luc"] & m["h_dkgd"]]           # VSD hủy, HNX còn ĐKGD
    D = only_vsd[only_vsd["hieu_luc"]]            # chỉ có ở VSD, còn hiệu lực
    E = only_hnx[only_hnx["dkgd"]]                # chỉ có ở HNX, đang ĐKGD

    v_hl, c_dk = vsd[vsd["hieu_luc"]], cat[cat["dkgd"]]
    n = lambda x: x / NGHIN_TY

    # ---- Sheet: Cầu nối
    bridge = pd.DataFrame([
        ["VSD — tổng giá trị mã còn HIỆU LỰC", len(v_hl), n(v_hl["gt"].sum()),
         "Điểm xuất phát: tổng giá trị CK đăng ký lưu ký của mã còn hiệu lực"],
        ["A. Cả 2 nguồn còn hiệu lực", len(A), n(A["gt"].sum()),
         f"HNX ghi nhận {n(A['h_gt'].sum()):,.1f} nghìn tỷ — chênh chỉ "
         f"{n(A['gt'].sum() - A['h_gt'].sum()):+,.2f} nghìn tỷ "
         f"({(A['gt'] == A['h_gt']).sum()}/{len(A)} mã khớp CHÍNH XÁC)"],
        ["B. (−) VSD hiệu lực nhưng HNX đã hủy ĐKGD", len(B), n(B["gt"].sum()),
         "Còn lưu ký ở VSD nhưng đã rời sàn TPRL → HNX tính KL ĐKGD = 0"],
        ["D. (−) Chỉ có ở VSD, không có trong danh mục ĐKGD", len(D), n(D["gt"].sum()),
         "Đã lưu ký nhưng chưa/không đăng ký giao dịch (TP mới 2026 + TP cũ trước 2021)"],
        ["C. (+) VSD đã hủy nhưng HNX còn ĐKGD", len(C), n(C["gt"].sum()), "—"],
        ["E. (+) Chỉ có ở HNX, đang ĐKGD", len(E), n(E["gt"].sum()),
         "ĐKGD trên HNX nhưng không thấy trong danh sách VSD"],
        ["HNX — tổng giá trị mã đang ĐKGD", len(c_dk), n(c_dk["gt"].sum()),
         "Điểm đến: KL ĐKGD × mệnh giá"],
    ], columns=["Cấu phần", "Số mã", "Giá trị (nghìn tỷ)", "Diễn giải"])

    # ---- Sheet: lệch số lượng khi CẢ HAI còn hiệu lực (khác biệt thực sự đáng ngờ)
    a = A.copy()
    a["chenh"] = a["gt"] - a["h_gt"]
    qa = a[a["chenh"] != 0].copy()
    qa["nghi_van"] = [
        "VSD > cả lượng phát hành CBTT" if (g > p > 0) else
        ("Chênh = mua lại luỹ kế" if abs(-c - ml) < TY and c < 0 else "Cần rà soát")
        for g, p, c, ml in zip(qa["gt"], qa["gt_ph"], qa["chenh"], qa["gt_ml"])]
    q1 = pd.DataFrame({
        "Mã CK (GD)": qa["ma_ck"], "Mã CBTT": qa["h_ma_cbtt"], "ISIN": qa["isin"],
        "Tên TCPH": qa["ten_tcdkck"],
        "Số lượng VSD": qa["sl"], "KL ĐKGD HNX": qa["h_kl"],
        "Mệnh giá": qa["mg"],
        "Giá trị VSD (tỷ)": qa["gt"] / TY, "Giá trị HNX (tỷ)": qa["h_gt"] / TY,
        "Chênh giá trị (tỷ)": qa["chenh"] / TY,
        "GT phát hành CBTT (tỷ)": qa["gt_ph"] / TY,
        "Mua lại luỹ kế (tỷ)": qa["gt_ml"] / TY,
        "Nghi vấn": qa["nghi_van"],
    }).sort_values("Chênh giá trị (tỷ)", key=abs, ascending=False)

    # ---- Sheet: lệch trạng thái
    st = pd.DataFrame({
        "Mã CK (GD)": m["ma_ck"], "Mã CBTT": m["h_ma_cbtt"], "ISIN": m["isin"],
        "Tên TCPH": m["ten_tcdkck"], "Tình trạng VSD": m["tinh_trang"],
        "Trạng thái HNX": m["h_trang_thai"], "Ngày đáo hạn": m["ngay_dao_han"],
        "Giá trị VSD (tỷ)": m["gt"] / TY, "Đã đáo hạn?": ["Đã đáo hạn" if _matured(d) else "Chưa đáo hạn" for d in m["d_dh"]],
        "Nối theo": m["join_how"],
    })[m["hieu_luc"] != m["h_dkgd"]].sort_values("Giá trị VSD (tỷ)", ascending=False)

    # ---- Sheet: lệch ngày PH/ĐH so với CBTT phát hành HNX
    dd = m[m["i_ph"].notna()].copy()
    dd["lech_ph"] = [(x - y).days if isinstance(x, date) and isinstance(y, date) else None
                     for x, y in zip(dd["d_ph"], dd["i_ph"])]
    dd["lech_dh"] = [(x - y).days if isinstance(x, date) and isinstance(y, date) else None
                     for x, y in zip(dd["d_dh"], dd["i_dh"])]
    dd = dd[(dd["lech_ph"].fillna(0) != 0) | (dd["lech_dh"].fillna(0) != 0)]
    dcmp = pd.DataFrame({
        "Mã CK (GD)": dd["ma_ck"], "Mã CBTT": dd["h_ma_cbtt"], "Tên TCPH": dd["ten_tcdkck"],
        "Ngày PH (VSD)": dd["ngay_phat_hanh"], "Ngày PH (HNX CBTT)": dd["i_ph"].map(_dstr),
        "Lệch PH (ngày)": dd["lech_ph"],
        "Ngày ĐH (VSD)": dd["ngay_dao_han"], "Ngày ĐH (HNX CBTT)": dd["i_dh"].map(_dstr),
        "Lệch ĐH (ngày)": dd["lech_dh"], "Số đợt CBTT": dd["n_dot"],
    }).sort_values("Lệch ĐH (ngày)", key=lambda s: s.abs(), ascending=False, na_position="last")

    # ---- độ phủ CBTT: thử CẢ mã CBTT LẪN mã GD (mã mới: CBTT == mã GD)
    cov = set(m["h_ma_cbtt"]) | set(vsd["cbtt_k"]) | set(vsd["gd_k"])
    cov.discard("")
    iss_in = iss["cbtt_k"].isin(cov)
    mis = iss[~iss_in].copy()
    mis["da_dh"] = [_matured(d) for d in mis["i_dh"]]

    # ---- Sheet: chỉ có ở VSD
    ov_key = only_vsd["cbtt_k"].where(only_vsd["cbtt_k"] != "", only_vsd["gd_k"])
    ov = pd.DataFrame({
        "Mã CK (GD)": only_vsd["ma_ck"], "Mã CBTT": only_vsd["ma_cbtt"], "ISIN": only_vsd["isin"],
        "Tên TCĐKCK": only_vsd["ten_tcdkck"], "Kỳ hạn": only_vsd["ky_han"],
        "Ngày phát hành": only_vsd["ngay_phat_hanh"], "Ngày đáo hạn": only_vsd["ngay_dao_han"],
        "Số lượng": only_vsd["sl"], "Tình trạng VSD": only_vsd["tinh_trang"],
        "Giá trị ĐK (tỷ)": only_vsd["gt"] / TY,
        "Năm PH": [d.year if d else None for d in only_vsd["d_ph"]],
        "Đối chiếu CBTT HNX": [
            "Có ở CBTT phát hành HNX (chưa ĐKGD)" if k in set(iss["cbtt_k"])
            else "Không có ở HNX" for k in ov_key],
    }).sort_values("Giá trị ĐK (tỷ)", ascending=False)

    # ---- Sheet: chỉ có ở HNX ĐKGD
    oh = pd.DataFrame({
        "Mã CBTT": only_hnx["ma_cbtt"], "Mã GD": only_hnx["ma_gd"], "ISIN": only_hnx["isin"],
        "Tên TCPH": only_hnx["ten_tcph"], "KL ĐKGD": only_hnx["kl"],
        "Mệnh giá": only_hnx["mg"], "Trạng thái ĐKGD": only_hnx["trang_thai"],
        "Ngày GD đầu": only_hnx["ngay_gd_dau"], "Ngày GD cuối": only_hnx["ngay_gd_cuoi"],
        "Giá trị (tỷ)": only_hnx["gt"] / TY,
    }).sort_values("Giá trị (tỷ)", ascending=False)

    # ---- Sheet: CBTT HNX không tìm thấy ở VSD
    # Phân loại theo dư nợ còn lại: mã đã đáo hạn hoặc đã mua lại hết thì VSD gỡ là ĐÚNG;
    # chỉ mã CHƯA đáo hạn & CÒN dư nợ mới thực sự là khoảng trống dữ liệu.
    mis = mis.merge(bb, on="cbtt_k", how="left")
    mis["gt_ml"] = mis["gt_ml"].fillna(0)
    mis["du_no"] = (mis["gt_ph"] - mis["gt_ml"]).clip(lower=0)
    mis["phan_loai"] = [
        "Đã đáo hạn — VSD gỡ khỏi danh sách (bình thường)" if da else
        ("Đã mua lại hết trước hạn (bình thường)" if dn <= 0 else
         "CHƯA đáo hạn & CÒN dư nợ — CẦN RÀ SOÁT")
        for da, dn in zip(mis["da_dh"], mis["du_no"])]
    nv = pd.DataFrame({
        "Mã CBTT": mis["cbtt_k"], "Tên DN": mis["ten_dn"],
        "Ngày phát hành": mis["i_ph"].map(_dstr), "Ngày đáo hạn": mis["i_dh"].map(_dstr),
        "Số đợt": mis["n_dot"], "Giá trị phát hành (tỷ)": mis["gt_ph"] / TY,
        "Mua lại luỹ kế (tỷ)": mis["gt_ml"] / TY, "Dư nợ còn lại (tỷ)": mis["du_no"] / TY,
        "Phân loại": mis["phan_loai"],
    }).sort_values(["Phân loại", "Dư nợ còn lại (tỷ)"], ascending=[True, False])
    n_soat = int((mis["phan_loai"].str.startswith("CHƯA")).sum())
    gt_soat = mis.loc[mis["phan_loai"].str.startswith("CHƯA"), "du_no"].sum()

    # ---- Sheet: tổng quan
    n_ex = int((A["gt"] == A["h_gt"]).sum())
    rows = [
        ["PHẠM VI DỮ LIỆU", "", "", ""],
        ["Số mã TPDN riêng lẻ", len(vsd), len(cat),
         "VSD = đăng ký lưu ký tại VSDC | HNX = danh mục đăng ký giao dịch sàn TPRL"],
        ["  · còn hiệu lực / đang ĐKGD", len(v_hl), len(c_dk), "VSD 'Hiệu lực' vs HNX 'Đăng ký giao dịch'"],
        ["  · đã hủy đăng ký", len(vsd) - len(v_hl), len(cat) - len(c_dk), ""],
        ["Giá trị mã còn hiệu lực (nghìn tỷ)", n(v_hl["gt"].sum()), n(c_dk["gt"].sum()),
         "Chênh giải thích đầy đủ ở sheet 'Cầu nối chênh lệch'"],
        ["", "", "", ""],
        ["ĐỘ PHỦ KHI NỐI KHÓA", "", "", ""],
        ["Mã khớp được 2 nguồn", len(matched), len(matched),
         f"ISIN {int((matched['join_how'] == 'ISIN').sum())} · "
         f"mã CBTT {int((matched['join_how'] == 'Mã CBTT').sum())} · "
         f"mã GD {int((matched['join_how'] == 'Mã GD').sum())}"],
        ["Chỉ có ở VSD", len(only_vsd), "",
         f"{int((ov['Đối chiếu CBTT HNX'].str.startswith('Có')).sum())} mã có ở CBTT phát hành HNX "
         f"nhưng chưa ĐKGD; còn lại là TP cũ trước thời kỳ ĐKGD"],
        ["Chỉ có ở HNX (ĐKGD)", "", len(only_hnx), "ĐKGD nhưng không thấy ở danh sách VSD"],
        ["", "", "", ""],
        ["MỨC ĐỘ ĐỒNG THUẬN (mã cả 2 nguồn còn hiệu lực)", "", "", ""],
        ["Số mã đối chiếu được", len(A), len(A), "nhóm A của cầu nối"],
        ["  · khớp CHÍNH XÁC giá trị", n_ex, n_ex,
         f"{n_ex/len(A)*100:.1f}% số mã — hai nguồn độc lập cho cùng con số tới từng đồng"],
        ["  · lệch giá trị", len(A) - n_ex, len(A) - n_ex, "chi tiết ở sheet 'Lệch SL-GT (cùng hiệu lực)'"],
        ["  · lệch mệnh giá", int((A["mg"] != A["h_mg"]).sum()), "", "0 = mệnh giá hoàn toàn thống nhất"],
        ["Tổng giá trị nhóm A (nghìn tỷ)", n(A["gt"].sum()), n(A["h_gt"].sum()),
         f"lệch {n(A['gt'].sum() - A['h_gt'].sum()):+,.2f} nghìn tỷ = "
         f"{(A['gt'].sum() - A['h_gt'].sum()) / A['h_gt'].sum() * 100:+.2f}%"],
        ["", "", "", ""],
        ["CHÊNH LỆCH KHÁC", "", "", ""],
        ["Lệch trạng thái", len(st), "", "hiệu lực ở nguồn này nhưng đã hủy ở nguồn kia"],
        ["Lệch ngày PH/ĐH", len(dcmp), "", "VSD so với CBTT phát hành HNX"],
        ["", "", "", ""],
        ["ĐỐI CHIẾU VỚI CBTT PHÁT HÀNH HNX", "", "", ""],
        ["Mã ở CBTT phát hành HNX", "", len(iss), "bond_issuance_raw (dữ liệu từ ~30/12/2020)"],
        ["  · tìm thấy ở VSD", int(iss_in.sum()), "", "nối theo mã CBTT hoặc mã GD"],
        ["  · KHÔNG thấy ở VSD", int((~iss_in).sum()), "",
         f"{int(mis['da_dh'].sum())} mã đã đáo hạn + "
         f"{int((~mis['da_dh']).sum()) - n_soat} mã đã mua lại hết → VSD gỡ là ĐÚNG"],
        ["  · → thực sự là khoảng trống", n_soat, "",
         f"chưa đáo hạn & còn dư nợ {gt_soat / NGHIN_TY:,.1f} nghìn tỷ theo HNX "
         f"nhưng không có ở VSD — xem sheet 'CBTT HNX thiếu ở VSD'"],
    ]
    ovw = pd.DataFrame(rows, columns=["Chỉ tiêu", "VSD", "HNX", "Ghi chú"])

    # ---- xuất Excel
    wb = Workbook()
    wb.remove(wb.active)
    stamp = f"Nguồn: vsd.vn/vi/ibl & cbonds.hnx.vn — lập lúc {datetime.now():%d/%m/%Y %H:%M}"
    write_sheet(wb, "Tổng quan", ovw, "ĐỐI CHIẾU TPDN RIÊNG LẺ: VSD vs HNX. " + stamp,
                widths={"Chỉ tiêu": 46, "VSD": 14, "HNX": 14, "Ghi chú": 86})
    write_sheet(wb, "Cầu nối chênh lệch", bridge,
                "Giải thích TRỌN VẸN chênh lệch giữa 'VSD hiệu lực' và 'HNX đang ĐKGD': "
                "A + B + D = VSD ; A + C + E = HNX.",
                widths={"Cấu phần": 48, "Số mã": 9, "Giá trị (nghìn tỷ)": 18, "Diễn giải": 78})
    write_sheet(wb, "Lệch SL-GT (cùng hiệu lực)", q1,
                "CHỈ gồm mã mà CẢ HAI nguồn đều ghi còn hiệu lực → chênh ở đây là bất thường dữ liệu thật, "
                "không phải khác biệt định nghĩa. Cột 'GT phát hành CBTT' để soi ngược mã nào vô lý.")
    write_sheet(wb, "Lệch trạng thái", st,
                "VSD 'Hiệu lực/Hủy đăng ký' vs HNX 'Đăng ký giao dịch/Đã hủy ĐKGD'. "
                "Hủy ĐKGD (rời sàn) KHÔNG đồng nghĩa hủy lưu ký.")
    write_sheet(wb, "Lệch ngày PH-ĐH", dcmp,
                "Ngày phát hành/đáo hạn ở VSD so với CBTT phát hành HNX. "
                "Mã nhiều đợt (Số đợt > 1) lệch là bình thường do CBTT gộp nhiều đợt.")
    write_sheet(wb, "Chỉ có ở VSD", ov,
                "Có ở VSD nhưng không nối được sang danh mục ĐKGD HNX "
                "(đã lưu ký nhưng chưa/không đăng ký giao dịch).")
    write_sheet(wb, "Chỉ có ở HNX (ĐKGD)", oh,
                "Có trong danh mục ĐKGD HNX nhưng không thấy trong danh sách VSD.")
    write_sheet(wb, "CBTT HNX thiếu ở VSD", nv,
                "Đã CBTT phát hành trên HNX nhưng không tìm thấy ở VSD (nối theo mã CBTT hoặc mã GD). "
                "Mã đã đáo hạn là bình thường; mã CHƯA đáo hạn mới đáng rà soát.")
    wb.save(OUT_XLSX)

    # ---- tóm tắt màn hình
    print("\n" + "=" * 80)
    for _, r in ovw.iterrows():
        a_, b_ = r["VSD"], r["HNX"]
        fmt = lambda x: (f"{x:,.1f}" if isinstance(x, float)
                         else (f"{x:,}" if isinstance(x, int) else str(x)))
        if not r["Chỉ tiêu"].strip():
            continue
        if a_ == "" and b_ == "":
            print(f"\n{r['Chỉ tiêu']}")
        else:
            print(f"  {r['Chỉ tiêu']:<44} VSD={fmt(a_):>10}  HNX={fmt(b_):>10}")
    print("\n" + "-" * 80)
    print("CẦU NỐI (nghìn tỷ):")
    for _, r in bridge.iterrows():
        print(f"  {r['Cấu phần']:<50} {r['Số mã']:>5} mã  {r['Giá trị (nghìn tỷ)']:>8.1f}")
    print("=" * 80)
    print(f"\nĐã lưu -> {OUT_XLSX}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
